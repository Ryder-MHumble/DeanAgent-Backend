"""导入共建导师协议信息并通过 AMiner 富化学者数据。

流程：
  1. 读取 Excel，解析 140 条导师记录
  2. 遍历 data/scholars/ 下所有 latest.json，建立姓名+单位索引
  3. 对每条导师记录：
     a. 按姓名+单位匹配已有学者（精确匹配优先，单位宽松匹配次之）
     b. 将 adjunct_supervisor 协议信息写入 annotation store
     c. 调用 AMiner person/search + person/detail 富化学者基本信息
        - 匹配置信度高时更新：bio, email, h_index, citations_count,
          publications_count, name_en, photo_url, google_scholar_url
     d. 未匹配到已有记录时：创建新记录存入
        data/scholars/adjunct_import/adjunct_supervisors/latest.json
  4. 打印导入汇总

用法：
  python scripts/data_import/import_adjunct_supervisors.py
  python scripts/data_import/import_adjunct_supervisors.py --dry-run
  python scripts/data_import/import_adjunct_supervisors.py --skip-aminer
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import re
import unicodedata
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parents[2]
EXCEL_PATH = ROOT / "docs" / "共建导师协议签署情况表20260228update.xlsx"
SCHOLARS_DIR = ROOT / "data" / "scholars"
ANNOTATIONS_FILE = ROOT / "data" / "state" / "scholar_annotations.json"
IMPORT_DIR = SCHOLARS_DIR / "adjunct_import" / "adjunct_supervisors"

# ---------------------------------------------------------------------------
# Institutions: uni_name_zh → AMiner org_name
# ---------------------------------------------------------------------------

_UNI_ORG_MAP: dict[str, str] = {
    "清华大学": "Tsinghua University",
    "北京大学": "Peking University",
    "中国人民大学": "Renmin University of China",
    "北京理工大学": "Beijing Institute of Technology",
    "北京航空航天大学": "Beihang University",
    "北京邮电大学": "Beijing University of Posts and Telecommunications",
    "北京师范大学": "Beijing Normal University",
    "武汉大学": "Wuhan University",
    "上海交通大学": "Shanghai Jiao Tong University",
    "浙江大学": "Zhejiang University",
    "复旦大学": "Fudan University",
    "南京大学": "Nanjing University",
    "中国科学技术大学": "University of Science and Technology of China",
    "南开大学": "Nankai University",
    "哈尔滨工业大学": "Harbin Institute of Technology",
    "南方科技大学": "Southern University of Science and Technology",
    "中山大学": "Sun Yat-sen University",
    "天津大学": "Tianjin University",
    "厦门大学": "Xiamen University",
    "吉林大学": "Jilin University",
    "中国科学院自动化所": "Institute of Automation, Chinese Academy of Sciences",
    "中科院自动化所": "Institute of Automation, Chinese Academy of Sciences",
    "中国科学院自动化研究所": "Institute of Automation, Chinese Academy of Sciences",
    "中国科学院大学": "University of Chinese Academy of Sciences",
    "西湖大学": "Westlake University",
    "香港科技大学": "Hong Kong University of Science and Technology",
    "华南师范大学": "South China Normal University",
    "深圳大学": "Shenzhen University",
    "中央财经大学": "Central University of Finance and Economics",
    "中国政法大学": "China University of Political Science and Law",
    "中央民族大学": "Minzu University of China",
    "中国农业大学": "China Agricultural University",
    "深圳湾实验室": "Shenzhen Bay Laboratory",
    "昌平国家实验室": "Changping National Laboratory",
    "上海AI Lab": "Shanghai AI Laboratory",
    "深圳医学科学院": "Shenzhen Medical Academy of Research and Translation",
    "山东大学": "Shandong University",
    "西安交通大学": "Xi'an Jiaotong University",
    "电子科技大学": "University of Electronic Science and Technology of China",
}


def _get_org_name(uni_raw: str | None) -> str:
    """Map Chinese university name to AMiner org_name (best effort)."""
    if not uni_raw:
        return ""
    uni_raw = uni_raw.strip()
    # Exact match
    if uni_raw in _UNI_ORG_MAP:
        return _UNI_ORG_MAP[uni_raw]
    # Prefix match (handles "北京理工大学自动化学院" etc.)
    for zh, en in _UNI_ORG_MAP.items():
        if uni_raw.startswith(zh) or zh in uni_raw:
            return en
    return ""


# ---------------------------------------------------------------------------
# Text normalization for matching
# ---------------------------------------------------------------------------

def _normalize(s: str) -> str:
    """Normalize for fuzzy matching: strip, lower, remove spaces/punctuation."""
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\s\u3000\-_()（）【】]", "", s)
    return s.lower()


def _uni_key(raw: str | None) -> str:
    """Extract the core university name for matching."""
    if not raw:
        return ""
    raw = raw.strip()
    # Strip department suffixes: "北京理工大学自动化学院" → "北京理工大学"
    for zh in sorted(_UNI_ORG_MAP.keys(), key=len, reverse=True):
        if raw.startswith(zh):
            return _normalize(zh)
    return _normalize(raw)


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def load_excel(path: Path) -> list[dict[str, Any]]:
    """Parse the adjunct supervisor Excel into a list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]  # row 2 is actual header row

    col = {h: i for i, h in enumerate(headers) if h}

    records: list[dict[str, Any]] = []
    for row in rows[2:]:
        name = row[col["导师姓名"]] if "导师姓名" in col else None
        uni = row[col["就职单位"]] if "就职单位" in col else None

        if not name or not uni or str(uni).strip() in ("——", "—", ""):
            continue
        name = str(name).strip()
        uni = str(uni).strip()
        if not name or name in ("——", "—"):
            continue

        records.append({
            "name": name,
            "university": uni,
            "position": str(row[col["职称"]]).strip() if "职称" in col and row[col["职称"]] else "",
            "research_areas_raw": str(row[col["研究领域"]]).strip() if "研究领域" in col and row[col["研究领域"]] else "",
            "phone": str(row[col["联系方式"]]).strip() if "联系方式" in col and row[col["联系方式"]] else "",
            "email": str(row[col["邮箱"]]).strip() if "邮箱" in col and row[col["邮箱"]] else "",
            "talent_program": str(row[col["人才计划"]]).strip() if "人才计划" in col and row[col["人才计划"]] else "",
            "adjunct_status": str(row[col["状态"]]).strip() if "状态" in col and row[col["状态"]] else "",
            "adjunct_type": str(row[col["类型"]]).strip() if "类型" in col and row[col["类型"]] else "",
            "agreement_type": str(row[col["协议类型"]]).strip() if "协议类型" in col and row[col["协议类型"]] else "",
            "agreement_period": str(row[col["协议签署时间"]]).strip() if "协议签署时间" in col and row[col["协议签署时间"]] else "",
            "recommender": str(row[col["推荐主体"]]).strip() if "推荐主体" in col and row[col["推荐主体"]] else "",
        })
    return records


# ---------------------------------------------------------------------------
# Scholar JSON index
# ---------------------------------------------------------------------------

def build_scholar_index() -> dict[str, dict[str, Any]]:
    """Build index: normalized_name → {url_hash, file_path, item_idx, name, university}."""
    index: dict[str, list[dict[str, Any]]] = {}
    for json_path in SCHOLARS_DIR.rglob("latest.json"):
        try:
            with open(json_path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue
        for idx, item in enumerate(data.get("items", [])):
            extra = item.get("extra") or {}
            name = extra.get("name", "") or item.get("title", "")
            uni = extra.get("university", "")
            if not name:
                continue
            key = _normalize(name)
            entry = {
                "url_hash": item.get("url_hash", ""),
                "file_path": str(json_path),
                "item_idx": idx,
                "name": name,
                "university": uni,
                "uni_key": _uni_key(uni),
            }
            index.setdefault(key, []).append(entry)
    return index


def find_scholar(
    name: str,
    university: str,
    index: dict[str, list[dict[str, Any]]],
) -> dict[str, Any] | None:
    """Find best matching scholar entry for (name, university)."""
    name_key = _normalize(name)
    candidates = index.get(name_key, [])
    if not candidates:
        return None

    target_uni_key = _uni_key(university)

    # Exact university match
    for c in candidates:
        if c["uni_key"] == target_uni_key:
            return c

    # Partial university match (one contains the other)
    for c in candidates:
        ck = c["uni_key"]
        if ck and target_uni_key and (ck in target_uni_key or target_uni_key in ck):
            return c

    # If only one candidate, return it (different departments / name variants)
    if len(candidates) == 1:
        return candidates[0]

    return None


# ---------------------------------------------------------------------------
# Annotation store helpers (no dependency on FastAPI app)
# ---------------------------------------------------------------------------

def _load_annotations() -> dict[str, Any]:
    if ANNOTATIONS_FILE.exists():
        with open(ANNOTATIONS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_annotations(data: dict[str, Any]) -> None:
    ANNOTATIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(ANNOTATIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def write_adjunct_annotation(
    url_hash: str,
    adjunct_info: dict[str, str],
    dry_run: bool = False,
) -> None:
    """Write adjunct_supervisor annotation for a scholar."""
    if dry_run:
        return
    data = _load_annotations()
    ann = data.setdefault(url_hash, {})
    ann["adjunct_supervisor"] = adjunct_info
    ann["relation_updated_at"] = datetime.now(UTC).isoformat()
    ann["relation_updated_by"] = "excel_import"
    _save_annotations(data)


# ---------------------------------------------------------------------------
# AMiner enrichment
# ---------------------------------------------------------------------------

async def aminer_search(name: str, org: str, api_key: str) -> list[dict]:
    """Call AMiner person/search, return list of candidates."""
    import httpx
    url = "https://datacenter.aminer.cn/gateway/open_platform/api/person/search"
    payload = {"name": name, "org": org, "size": 5, "offset": 0}
    headers = {"Content-Type": "application/json;charset=utf-8", "Authorization": api_key}
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.post(url, json=payload, headers=headers)
        r.raise_for_status()
        return r.json().get("data", [])


async def aminer_detail(aminer_id: str, api_key: str) -> dict:
    """Call AMiner person/detail, return raw data dict."""
    import httpx
    url = "https://datacenter.aminer.cn/gateway/open_platform/api/person/detail"
    headers = {"Content-Type": "application/json;charset=utf-8", "Authorization": api_key}
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.get(url, params={"id": aminer_id}, headers=headers)
        r.raise_for_status()
        return r.json().get("data", {})


def _pick_best_aminer_candidate(
    candidates: list[dict],
    name: str,
    university: str,
) -> dict | None:
    """Return the AMiner candidate that best matches name + university."""
    name_norm = _normalize(name)
    uni_key = _uni_key(university)

    scored: list[tuple[int, dict]] = []
    for c in candidates:
        score = 0
        # Name match
        cn = _normalize(c.get("name", "") + c.get("name_zh", ""))
        if name_norm in cn or cn == name_norm:
            score += 2
        # Org match
        org_text = _normalize(c.get("org", ""))
        if uni_key and uni_key in org_text:
            score += 3
        scored.append((score, c))

    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best = scored[0] if scored else (0, None)
    # Require at least name match to avoid false positives
    return best if best_score >= 2 else None


def _map_aminer_to_scholar(detail: dict) -> dict[str, Any]:
    """Map AMiner person/detail response to ScholarRecord field updates."""
    updates: dict[str, Any] = {}

    if detail.get("name"):
        updates["name_en"] = detail["name"]
    if detail.get("avatar"):
        updates["photo_url"] = detail["avatar"]
    if detail.get("bio"):
        updates["bio"] = detail["bio"]
    if detail.get("email"):
        updates["email"] = detail["email"]
    if detail.get("homepage"):
        updates["google_scholar_url"] = detail["homepage"]

    indices = detail.get("indices") or {}
    if indices.get("hindex", -1) >= 0:
        updates["h_index"] = indices["hindex"]
    if indices.get("citations", -1) >= 0:
        updates["citations_count"] = indices["citations"]
    if indices.get("pubs", -1) >= 0:
        updates["publications_count"] = indices["pubs"]

    # Education
    edu_list = detail.get("educations") or []
    if edu_list:
        education = []
        for e in edu_list:
            education.append({
                "degree": e.get("degree", ""),
                "institution": e.get("institution", "") or e.get("org", ""),
                "year": str(e.get("year", "") or ""),
                "major": e.get("major", ""),
            })
        if education:
            updates["education"] = education
            # Populate phd shortcut
            phd = next((e for e in education if "博士" in e.get("degree", "") or "Ph" in e.get("degree", "")), None)
            if phd:
                updates["phd_institution"] = phd["institution"]
                updates["phd_year"] = phd["year"]

    updates["metrics_updated_at"] = datetime.now(UTC).isoformat()
    return updates


def update_scholar_json(
    file_path: str,
    item_idx: int,
    field_updates: dict[str, Any],
    dry_run: bool = False,
) -> None:
    """Apply field_updates to item's extra dict in-place and save."""
    if dry_run:
        return
    path = Path(file_path)
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    extra = data["items"][item_idx].get("extra") or {}
    for k, v in field_updates.items():
        # Only overwrite if target field is empty / default
        current = extra.get(k)
        if current in (None, "", [], -1, False):
            extra[k] = v
    data["items"][item_idx]["extra"] = extra
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# Create new scholar for unmatched records
# ---------------------------------------------------------------------------

def _make_url_hash(name: str, university: str) -> str:
    key = f"adjunct_import:{name}:{university}"
    return hashlib.sha256(key.encode()).hexdigest()


def create_new_scholar_record(rec: dict[str, Any]) -> dict[str, Any]:
    """Build a minimal CrawledItem dict from an Excel row."""
    now = datetime.now(UTC).isoformat()
    name = rec["name"]
    uni = rec["university"]
    url_hash = _make_url_hash(name, uni)
    research_areas = [a.strip() for a in re.split(r"[；;、，,/\\|\n\r]+", rec["research_areas_raw"]) if a.strip()] if rec["research_areas_raw"] else []
    academic_titles = [rec["talent_program"]] if rec["talent_program"] else []

    extra = {
        "name": name,
        "name_en": "",
        "gender": "",
        "photo_url": "",
        "university": uni,
        "department": "",
        "secondary_departments": [],
        "position": rec["position"],
        "academic_titles": academic_titles,
        "is_academician": False,
        "research_areas": research_areas,
        "keywords": [],
        "bio": "",
        "bio_en": "",
        "email": rec["email"],
        "phone": rec["phone"],
        "office": "",
        "profile_url": "",
        "lab_url": "",
        "google_scholar_url": "",
        "dblp_url": "",
        "orcid": "",
        "phd_institution": "",
        "phd_year": "",
        "education": [],
        "publications_count": -1,
        "h_index": -1,
        "citations_count": -1,
        "metrics_updated_at": "",
        "representative_publications": [],
        "patents": [],
        "awards": [],
        "is_advisor_committee": False,
        "adjunct_supervisor": {
            "status": rec["adjunct_status"],
            "type": rec["adjunct_type"],
            "agreement_type": rec["agreement_type"],
            "agreement_period": rec["agreement_period"],
            "recommender": rec["recommender"],
        },
        "supervised_students": [],
        "joint_research_projects": [],
        "joint_management_roles": [],
        "academic_exchange_records": [],
        "is_potential_recruit": False,
        "institute_relation_notes": "",
        "relation_updated_by": "excel_import",
        "relation_updated_at": now,
        "recent_updates": [],
        "source_id": "adjunct_supervisors",
        "source_url": "",
        "crawled_at": now,
        "first_seen_at": now,
        "last_seen_at": now,
        "is_active": True,
        "data_completeness": 0,
    }

    return {
        "title": name,
        "url": f"adjunct_import:{name}:{uni}",
        "url_hash": url_hash,
        "published_at": now,
        "author": "",
        "content": "",
        "content_html": "",
        "content_hash": "",
        "source_id": "adjunct_supervisors",
        "dimension": "scholars",
        "tags": [],
        "extra": extra,
        "is_new": True,
    }


def save_new_scholars(new_items: list[dict[str, Any]], dry_run: bool = False) -> None:
    """Write new scholar records to the adjunct_import source JSON."""
    if not new_items or dry_run:
        return
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = IMPORT_DIR / "latest.json"
    existing: list[dict] = []
    if out_path.exists():
        with open(out_path, encoding="utf-8") as f:
            existing = json.load(f).get("items", [])

    # Deduplicate by url_hash
    existing_hashes = {i["url_hash"] for i in existing}
    to_add = [i for i in new_items if i["url_hash"] not in existing_hashes]
    all_items = existing + to_add

    payload = {
        "source_id": "adjunct_supervisors",
        "source_name": "共建导师（Excel导入）",
        "group": "adjunct_import",
        "crawled_at": datetime.now(UTC).isoformat(),
        "items": all_items,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    logger.info("Saved %d new scholars to %s", len(to_add), out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def run(dry_run: bool, skip_aminer: bool) -> None:
    from app.config import settings
    api_key = settings.AMINER_API_KEY

    logger.info("Loading Excel: %s", EXCEL_PATH)
    records = load_excel(EXCEL_PATH)
    logger.info("Loaded %d valid records from Excel", len(records))

    logger.info("Building scholar index from data/scholars/...")
    scholar_index = build_scholar_index()
    logger.info("Indexed %d distinct scholar names", len(scholar_index))

    stats = {"matched": 0, "new": 0, "aminer_enriched": 0, "aminer_failed": 0, "skipped": 0}
    new_items: list[dict[str, Any]] = []

    for i, rec in enumerate(records):
        name = rec["name"]
        uni = rec["university"]
        logger.info("[%d/%d] Processing: %s @ %s", i + 1, len(records), name, uni)

        adjunct_info = {
            "status": rec["adjunct_status"],
            "type": rec["adjunct_type"],
            "agreement_type": rec["agreement_type"],
            "agreement_period": rec["agreement_period"],
            "recommender": rec["recommender"],
        }

        # ── Step 1: Match to existing scholar ─────────────────────────────
        matched = find_scholar(name, uni, scholar_index)

        if matched:
            stats["matched"] += 1
            url_hash = matched["url_hash"]
            logger.info("  ✓ Matched: %s (hash=%s...)", matched["name"], url_hash[:12])

            # Write adjunct annotation
            write_adjunct_annotation(url_hash, adjunct_info, dry_run=dry_run)

            # Also update existing extra fields from Excel if empty
            excel_updates: dict[str, Any] = {}
            if rec["position"]:
                excel_updates["position"] = rec["position"]
            if rec["email"]:
                excel_updates["email"] = rec["email"]
            if rec["phone"]:
                excel_updates["phone"] = rec["phone"]
            if rec["research_areas_raw"]:
                excel_updates["research_areas"] = [
                    a.strip() for a in re.split(r"[；;、，,/\\|\n\r]+", rec["research_areas_raw"]) if a.strip()
                ]
            if rec["talent_program"]:
                excel_updates["academic_titles"] = [rec["talent_program"]]
            if excel_updates:
                update_scholar_json(matched["file_path"], matched["item_idx"], excel_updates, dry_run=dry_run)

            # ── Step 2: AMiner enrichment ──────────────────────────────────
            if not skip_aminer and api_key:
                org_name = _get_org_name(uni)
                if not org_name:
                    logger.warning("  ⚠ No org_name for '%s', skipping AMiner", uni)
                else:
                    try:
                        candidates = await aminer_search(name, org_name, api_key)
                        best = _pick_best_aminer_candidate(candidates, name, uni)
                        if best:
                            detail = await aminer_detail(best["id"], api_key)
                            aminer_updates = _map_aminer_to_scholar(detail)
                            update_scholar_json(
                                matched["file_path"], matched["item_idx"],
                                aminer_updates, dry_run=dry_run,
                            )
                            stats["aminer_enriched"] += 1
                            logger.info(
                                "  ✓ AMiner enriched (id=%s, h=%s, cit=%s)",
                                best["id"], detail.get("indices", {}).get("hindex", "?"),
                                detail.get("indices", {}).get("citations", "?"),
                            )
                        else:
                            stats["aminer_failed"] += 1
                            logger.info("  ✗ AMiner: no confident match among %d candidates", len(candidates))
                    except Exception as e:
                        stats["aminer_failed"] += 1
                        logger.warning("  ✗ AMiner error for %s: %s", name, e)

        else:
            # ── Step 3: Create new record ──────────────────────────────────
            stats["new"] += 1
            logger.info("  ✗ No match found → will create new record")
            new_item = create_new_scholar_record(rec)

            # AMiner enrichment for new records too
            if not skip_aminer and api_key:
                org_name = _get_org_name(uni)
                if org_name:
                    try:
                        candidates = await aminer_search(name, org_name, api_key)
                        best = _pick_best_aminer_candidate(candidates, name, uni)
                        if best:
                            detail = await aminer_detail(best["id"], api_key)
                            aminer_updates = _map_aminer_to_scholar(detail)
                            # Merge into extra
                            for k, v in aminer_updates.items():
                                current = new_item["extra"].get(k)
                                if current in (None, "", [], -1, False):
                                    new_item["extra"][k] = v
                            stats["aminer_enriched"] += 1
                            logger.info("  ✓ AMiner enriched new record (id=%s)", best["id"])
                        else:
                            stats["aminer_failed"] += 1
                    except Exception as e:
                        stats["aminer_failed"] += 1
                        logger.warning("  ✗ AMiner error for new %s: %s", name, e)

            new_items.append(new_item)

    # Save new records
    save_new_scholars(new_items, dry_run=dry_run)

    # ── Summary ────────────────────────────────────────────────────────────
    print("\n" + "=" * 55)
    print(f"  导入完成{'（dry-run，未写入）' if dry_run else ''}")
    print("=" * 55)
    print(f"  总记录数:       {len(records)}")
    print(f"  匹配已有学者:   {stats['matched']}")
    print(f"  新建学者记录:   {stats['new']}")
    print(f"  AMiner 富化成功: {stats['aminer_enriched']}")
    print(f"  AMiner 未命中:  {stats['aminer_failed']}")
    print("=" * 55)


def main() -> None:
    parser = argparse.ArgumentParser(description="导入共建导师 Excel 数据")
    parser.add_argument("--dry-run", action="store_true", help="仅预览，不写入任何文件")
    parser.add_argument("--skip-aminer", action="store_true", help="跳过 AMiner API 调用")
    args = parser.parse_args()

    if args.dry_run:
        logger.info("DRY-RUN 模式：不会写入任何文件")

    asyncio.run(run(dry_run=args.dry_run, skip_aminer=args.skip_aminer))


if __name__ == "__main__":
    main()
