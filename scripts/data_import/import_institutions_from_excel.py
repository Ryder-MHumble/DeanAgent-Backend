#!/usr/bin/env python3
"""
Import university collaboration data from Excel to JSON storage.

Usage:
    python scripts/import_institutions_from_excel.py
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def parse_multiline_text(text: str | None) -> list[str]:
    """Parse multiline text into list of items."""
    if not text:
        return []
    lines = [line.strip() for line in str(text).split("\n") if line.strip()]
    return lines


def parse_person_info(text: str | None) -> list[dict]:
    """Parse person information from multiline text."""
    if not text:
        return []

    lines = parse_multiline_text(text)
    persons = []

    for line in lines:
        # Try to extract name and role
        parts = line.split("\n")
        if parts:
            persons.append({"name": parts[0].strip(), "info": line})

    return persons


def clean_priority(priority: str | None) -> tuple[str, bool]:
    """Clean priority field and extract demo school flag."""
    if not priority:
        return "", False

    priority_str = str(priority).strip()
    is_demo = "示范校" in priority_str

    # Extract P0/P1/P2/P3
    match = re.search(r"P[0-3]", priority_str)
    if match:
        return match.group(0), is_demo

    return priority_str, is_demo


def generate_institution_id(name: str) -> str:
    """Generate institution ID from name."""
    # Simple pinyin mapping for common universities
    mapping = {
        "清华大学": "tsinghua",
        "北京大学": "pku",
        "北京理工大学": "bit",
        "中国科学院大学": "ucas",
        "北京航空航天大学": "buaa",
        "北京邮电大学": "bupt",
        "北京师范大学": "bnu",
        "中国人民大学": "ruc",
        "哈尔滨工业大学": "hit",
        "上海交通大学": "sjtu",
    }

    if name in mapping:
        return mapping[name]

    # Fallback: use first characters
    return "".join([c for c in name if c.isalnum()])[:20].lower()


def main():
    excel_path = Path("docs/共建高校信息汇总.xlsx")
    output_dir = Path("data/institutions")
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Build headers from row 5 and 6
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        main_header = ws.cell(row=5, column=col_idx).value
        sub_header = ws.cell(row=6, column=col_idx).value

        if main_header and sub_header:
            header = f"{main_header}_{sub_header}"
        elif main_header:
            header = main_header
        elif sub_header:
            header = sub_header
        else:
            header = f"col_{col_idx}"

        headers.append(header)

    print(f"Found {len(headers)} columns")

    # Extract data starting from row 7
    institutions = []
    for row_idx in range(7, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[header] = cell.value

        # Skip empty rows
        university_name = row_data.get("高校")
        if not university_name:
            continue

        # Parse priority and demo school flag
        priority, is_demo = clean_priority(row_data.get("优先级"))

        # Build institution record
        institution_id = generate_institution_id(university_name)

        institution = {
            "id": institution_id,
            "name": university_name,
            "category": row_data.get("分类") or "",
            "priority": priority,
            "is_demo_school": is_demo,
            # Student counts
            "student_count_24": int(row_data.get("学生人数_24级") or 0),
            "student_count_25": int(row_data.get("25级") or 0),
            "student_count_total": int(row_data.get("总数") or 0),
            "supervisor_count": int(row_data.get("导师人数_总数") or 0),
            # Personnel
            "institute_leaders_raw": row_data.get("驻院领导及共建老师") or "",
            "degree_committee_raw": row_data.get("学位委员") or "",
            "teaching_committee_raw": row_data.get("教学委员") or "",
            "joint_supervisors_raw": row_data.get("共建导师_姓名") or "",
            "university_leaders_raw": row_data.get("相关校领导_姓名") or "",
            "important_scholars_raw": row_data.get("重要学者_姓名") or "",
            # Departments and collaboration
            "key_departments": parse_multiline_text(row_data.get("重点院系")),
            "collaboration_joint_lab": row_data.get("合作_联合实验室") or "",
            "collaboration_training": row_data.get("培养") or "",
            "collaboration_academic": row_data.get("学术") or "",
            "collaboration_talent_hiring": row_data.get("人才双聘") or "",
            "collaboration_focus": row_data.get("合作重点") or "",
            # Exchange records
            "exchange_recruitment": row_data.get("交流互访_招生宣讲") or "",
            "exchange_other": row_data.get("col_27") or "",
            # Metadata
            "data_source": "共建高校信息汇总.xlsx",
            "last_updated": "2025-07-18",
            "notes": "",
        }

        institutions.append(institution)
        print(f"  Processed: {university_name} ({institution_id})")

    # Save to JSON
    output_file = output_dir / "institutions.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(
            {
                "total": len(institutions),
                "last_updated": "2025-07-18",
                "institutions": institutions,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    print(f"\n✓ Imported {len(institutions)} institutions to {output_file}")

    # Print summary
    categories = {}
    priorities = {}
    for inst in institutions:
        cat = inst["category"] or "Unknown"
        pri = inst["priority"] or "Unknown"
        categories[cat] = categories.get(cat, 0) + 1
        priorities[pri] = priorities.get(pri, 0) + 1

    print("\nSummary:")
    print(f"  Total: {len(institutions)}")
    print(f"  Demo schools: {sum(1 for i in institutions if i['is_demo_school'])}")
    print(f"\n  By category:")
    for cat, count in sorted(categories.items()):
        print(f"    {cat}: {count}")
    print(f"\n  By priority:")
    for pri, count in sorted(priorities.items()):
        print(f"    {pri}: {count}")


if __name__ == "__main__":
    main()
