#!/usr/bin/env python3
"""
从 Excel 重建完整的 institutions.json。

功能：
1. 从 docs/共建高校信息汇总.xlsx 读取所有高校（41所）
2. 合并现有 institutions.json 中的学者数据（departments, scholar_count）
3. 生成完整的 institutions.json

输出结构：
{
  "last_updated": "...",
  "universities": [
    {
      "id": "...",
      "name": "...",
      "org_name": null,  # 后续通过 AMiner API 填充
      "scholar_count": 0,
      "category": "...",
      "priority": "...",
      "student_count_24": 0,
      "student_count_25": 0,
      "student_count_total": 0,
      "mentor_count": 0,
      "resident_leaders": [],
      "degree_committee": [],
      "teaching_committee": [],
      "mentors": [{name, category, department}],
      "university_leaders": ["name1", "name2"],
      "notable_scholars": ["name1", "name2"],
      "departments": [...]  # 从现有数据保留
    }
  ]
}
"""

import json
import openpyxl
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import re

DATA_DIR = Path(__file__).parent.parent / "data"
DOCS_DIR = Path(__file__).parent.parent / "docs"
INSTITUTIONS_FILE = DATA_DIR / "institutions.json"
EXCEL_FILE = DOCS_DIR / "共建高校信息汇总.xlsx"


def normalize_university_id(name: str) -> str:
    """将高校名称转换为 ID（拼音缩写）"""
    # 简单映射表
    name_to_id = {
        "清华大学": "tsinghua",
        "北京大学": "pku",
        "北京理工大学": "bit",
        "中国科学院大学": "ucas",
        "北京航空航天大学": "buaa",
        "北京邮电大学": "bupt",
        "北京师范大学": "bnu",
        "中国人民大学": "ruc",
        "哈尔滨工业大学": "hit",
        "上海交通大学": "sjtu",
        "中国科学技术大学": "ustc",
        "复旦大学": "fudan",
        "浙江大学": "zju",
        "西安交通大学": "xjtu",
        "南京大学": "nju",
        "同济大学": "tongji",
        "华中科技大学": "hust",
        "武汉大学": "whu",
        "东南大学": "seu",
        "中山大学": "sysu",
        "四川大学": "scu",
        "山东大学": "sdu",
        "厦门大学": "xmu",
        "天津大学": "tju",
        "南开大学": "nku",
        "大连理工大学": "dlut",
        "吉林大学": "jlu",
        "兰州大学": "lzu",
        "西北工业大学": "nwpu",
        "电子科技大学": "uestc",
        "重庆大学": "cqu",
        "湖南大学": "hnu",
        "中南大学": "csu",
        "华南理工大学": "scut",
        "北京交通大学": "bjtu",
        "北京科技大学": "ustb",
        "北京化工大学": "buct",
        "中国农业大学": "cau",
        "中国地质大学": "cug",
        "中国矿业大学": "cumt",
        "中国石油大学": "cup",
    }

    return name_to_id.get(name, name.lower().replace("大学", "").replace("学院", ""))


def load_existing_institutions() -> dict[str, dict]:
    """加载现有的 institutions.json，返回 {高校名: 高校数据} 字典"""
    if not INSTITUTIONS_FILE.exists():
        return {}

    with open(INSTITUTIONS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    existing = {}
    for uni in data.get("universities", []):
        existing[uni["name"]] = uni

    return existing


def parse_excel_all_universities() -> list[dict]:
    """解析 Excel 文件，返回所有高校列表"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # 表头在第5行，数据从第7行开始
    header_row = 5
    data_start_row = 7

    universities = []

    for i in range(data_start_row, ws.max_row + 1):
        # 读取基本信息
        category = ws.cell(i, 2).value  # 分类
        uni_name = ws.cell(i, 3).value  # 高校

        if not uni_name:
            continue

        uni_name = str(uni_name).strip()

        # 学生人数
        student_24 = ws.cell(i, 4).value
        student_25 = ws.cell(i, 5).value
        student_total = ws.cell(i, 6).value

        # 导师人数
        mentor_count = ws.cell(i, 7).value

        # 驻院领导及共建老师
        resident_leaders_raw = ws.cell(i, 8).value
        resident_leaders = []
        if resident_leaders_raw:
            resident_leaders = [x.strip() for x in str(resident_leaders_raw).split('\n') if x.strip()]

        # 学位委员
        degree_committee_raw = ws.cell(i, 9).value
        degree_committee = []
        if degree_committee_raw:
            degree_committee = [x.strip() for x in str(degree_committee_raw).split('\n') if x.strip()]

        # 教学委员
        teaching_committee_raw = ws.cell(i, 10).value
        teaching_committee = []
        if teaching_committee_raw:
            teaching_committee = [x.strip() for x in str(teaching_committee_raw).split('\n') if x.strip()]

        # 共建导师信息（姓名、类别、院系）
        mentor_names_raw = ws.cell(i, 11).value
        mentor_categories_raw = ws.cell(i, 12).value
        mentor_depts_raw = ws.cell(i, 13).value

        mentors = []
        if mentor_names_raw:
            names = [x.strip() for x in str(mentor_names_raw).split('\n') if x.strip()]
            categories = []
            if mentor_categories_raw:
                categories = [x.strip() for x in str(mentor_categories_raw).split('\n') if x.strip()]
            depts = []
            if mentor_depts_raw:
                depts = [x.strip() for x in str(mentor_depts_raw).split('\n') if x.strip()]

            for idx, name in enumerate(names):
                mentors.append({
                    "name": name,
                    "category": categories[idx] if idx < len(categories) else None,
                    "department": depts[idx] if idx < len(depts) else None
                })

        # 相关校领导
        leaders_raw = ws.cell(i, 14).value
        university_leaders = []
        if leaders_raw:
            university_leaders = [x.strip() for x in str(leaders_raw).split('\n') if x.strip()]

        # 重要学者（从第15列开始到最后一列之前）
        notable_scholars = []
        for j in range(15, ws.max_column):
            scholar_raw = ws.cell(i, j).value
            if scholar_raw:
                scholar_str = str(scholar_raw).strip()
                # 跳过优先级标记
                if not re.match(r'^P[0-3]', scholar_str):
                    notable_scholars.append(scholar_str)

        # 优先级（最后一列）
        priority_raw = ws.cell(i, ws.max_column).value
        priority = None
        if priority_raw:
            priority_str = str(priority_raw).strip()
            if 'P0' in priority_str:
                priority = 'P0'
            elif 'P1' in priority_str:
                priority = 'P1'
            elif 'P2' in priority_str:
                priority = 'P2'
            elif 'P3' in priority_str:
                priority = 'P3'

        universities.append({
            "id": normalize_university_id(uni_name),
            "name": uni_name,
            "org_name": None,  # 后续通过 AMiner API 填充
            "scholar_count": 0,  # 后续从现有数据合并
            "category": str(category).strip() if category else None,
            "priority": priority,
            "student_count_24": int(student_24) if student_24 and str(student_24).isdigit() else None,
            "student_count_25": int(student_25) if student_25 and str(student_25).isdigit() else None,
            "student_count_total": int(student_total) if student_total and str(student_total).isdigit() else None,
            "mentor_count": int(mentor_count) if mentor_count and str(mentor_count).isdigit() else None,
            "resident_leaders": resident_leaders,
            "degree_committee": degree_committee,
            "teaching_committee": teaching_committee,
            "mentors": mentors,
            "university_leaders": university_leaders,
            "notable_scholars": notable_scholars,
            "departments": [],  # 后续从现有数据合并
        })

    return universities


def merge_with_existing(new_universities: list[dict], existing: dict[str, dict]) -> list[dict]:
    """合并新高校列表和现有数据"""
    for uni in new_universities:
        uni_name = uni["name"]

        if uni_name in existing:
            # 合并现有数据
            existing_uni = existing[uni_name]

            # 保留现有的 departments, scholar_count, org_name
            uni["departments"] = existing_uni.get("departments", [])
            uni["scholar_count"] = existing_uni.get("scholar_count", 0)
            uni["org_name"] = existing_uni.get("org_name")

            print(f"  ✅ 合并现有数据: {uni_name} (学者: {uni['scholar_count']}, 院系: {len(uni['departments'])})")
        else:
            print(f"  🆕 新增高校: {uni_name}")

    return new_universities


def main():
    """主函数"""
    print("🚀 从 Excel 重建完整的 institutions.json\n")

    # 检查文件
    if not EXCEL_FILE.exists():
        print(f"❌ 文件不存在: {EXCEL_FILE}")
        return

    # 加载现有数据
    existing = load_existing_institutions()
    print(f"📊 现有高校数: {len(existing)}\n")

    # 解析 Excel
    new_universities = parse_excel_all_universities()
    print(f"\n📊 Excel 高校数: {len(new_universities)}\n")

    # 合并数据
    merged_universities = merge_with_existing(new_universities, existing)

    # 构建输出
    output = {
        "last_updated": datetime.now(timezone.utc).isoformat(),
        "universities": merged_universities
    }

    # 保存
    with open(INSTITUTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 完成！文件已保存: {INSTITUTIONS_FILE}")

    # 统计
    total_unis = len(merged_universities)
    total_depts = sum(len(u.get("departments", [])) for u in merged_universities)
    total_scholars = sum(u.get("scholar_count", 0) for u in merged_universities)
    unis_with_scholars = sum(1 for u in merged_universities if u.get("scholar_count", 0) > 0)

    print(f"\n📊 统计:")
    print(f"  总高校数: {total_unis}")
    print(f"  有学者数据的高校: {unis_with_scholars}")
    print(f"  总院系数: {total_depts}")
    print(f"  总学者数: {total_scholars}")


if __name__ == "__main__":
    main()
