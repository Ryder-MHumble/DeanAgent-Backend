#!/usr/bin/env python3
"""临时脚本：读取 Excel 文件表头"""
import sys
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

def read_excel_headers(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    print(f"\n{'='*60}")
    print(f"File: {filepath}")
    print(f"Total rows: {ws.max_row}, Total cols: {ws.max_column}")
    print(f"{'='*60}")

    # 读取前5行来理解结构
    print("\nFirst 5 rows:")
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row = [cell.value for cell in ws[row_idx]]
        non_empty = [v for v in row if v is not None]
        if non_empty:
            print(f"\nRow {row_idx}: {non_empty[:10]}")

if __name__ == "__main__":
    read_excel_headers("docs/讲座信息汇总.xlsx")
    read_excel_headers("docs/共建高校信息汇总.xlsx")
