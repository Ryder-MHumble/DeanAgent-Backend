#!/usr/bin/env python3
"""
构建完整的 institutions.json 文件。

数据来源：
1. 现有 institutions.json（高校/院系基本信息 + org_name）
2. docs/共建高校信息汇总.xlsx（院校领导、合作信息等详细信息）

输出：完整的 institutions.json，包含所有字段
"""

import json
import openpyxl
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

DATA_DIR = Path(__file__).parent.parent / "data"
DOCS_DIR = Path(__file__).parent.parent / "docs"
INSTITUTIONS_FILE = DATA_DIR / "institutions.json"
EXCEL_FILE = DOCS_DIR / "共建高校信息汇总.xlsx"


def load_existing_institutions() -> dict[str, Any]:
    """加载现有的 institutions.json（包含 org_name）"""
    if not INSTITUTIONS_FILE.exists():
        return {"universities": []}

    with open(INSTITUTIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_excel_data() -> dict[str, dict]:
    """解析 Excel 文件，返回 {高校名: 详细信息} 字典"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # 表头在第5行，数据从第7行开始
    header_row = 5
    data_start_row = 7

    # 读取表头
    headers = []
    for cell in ws[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip())

    # 读取第6行（子表头）
    sub_headers = []
    for cell in ws[6]:
        if cell.value:
            sub_headers.append(str(cell.value).strip())

    # 解析数据
    universities_data = {}

    for i in range(data_start_row, ws.max_row + 1):
        # 读取基本信息
        category = ws.cell(i, 2).value  # 分类
        uni_name = ws.cell(i, 3).value  # 高校

        if not uni_name:
            continue

        uni_name = str(uni_name).strip()

        # 学生人数
        student_24 = ws.cell(i, 4).value
        student_25 = ws.cell(i, 5).value
        student_total = ws.cell(i, 6).value

        # 导师人数
        mentor_count = ws.cell(i, 7).value

        # 驻院领导及共建老师
        resident_leaders_raw = ws.cell(i, 8).value
        resident_leaders = []
        if resident_leaders_raw:
            resident_leaders = [x.strip() for x in str(resident_leaders_raw).split('\n') if x.strip()]

        # 学位委员
        degree_committee_raw = ws.cell(i, 9).value
        degree_committee = []
        if degree_committee_raw:
            degree_committee = [x.strip() for x in str(degree_committee_raw).split('\n') if x.strip()]

        # 教学委员
        teaching_committee_raw = ws.cell(i, 10).value
        teaching_committee = []
        if teaching_committee_raw:
            teaching_committee = [x.strip() for x in str(teaching_committee_raw).split('\n') if x.strip()]

        # 共建导师信息（姓名、类别、院系）
        mentor_names_raw = ws.cell(i, 11).value
        mentor_categories_raw = ws.cell(i, 12).value
        mentor_depts_raw = ws.cell(i, 13).value

        mentors = []
        if mentor_names_raw:
            names = [x.strip() for x in str(mentor_names_raw).split('\n') if x.strip()]
            categories = []
            if mentor_categories_raw:
                categories = [x.strip() for x in str(mentor_categories_raw).split('\n') if x.strip()]
            depts = []
            if mentor_depts_raw:
                depts = [x.strip() for x in str(mentor_depts_raw).split('\n') if x.strip()]

            for idx, name in enumerate(names):
                mentors.append({
                    "name": name,
                    "category": categories[idx] if idx < len(categories) else None,
                    "department": depts[idx] if idx < len(depts) else None
                })

        # 相关校领导
        leaders_raw = ws.cell(i, 14).value
        university_leaders = []
        if leaders_raw:
            university_leaders = [x.strip() for x in str(leaders_raw).split('\n') if x.strip()]

        # 重要学者（从第15列开始）
        notable_scholars = []
        for j in range(15, ws.max_column + 1):
            scholar_raw = ws.cell(i, j).value
            if scholar_raw:
                notable_scholars.append(str(scholar_raw).strip())

        # 优先级（最后一列）
        priority_raw = ws.cell(i, ws.max_column).value
        priority = None
        if priority_raw:
            priority_str = str(priority_raw).strip()
            if 'P0' in priority_str:
                priority = 'P0'
            elif 'P1' in priority_str:
                priority = 'P1'
            elif 'P2' in priority_str:
                priority = 'P2'
            elif 'P3' in priority_str:
                priority = 'P3'

        universities_data[uni_name] = {
            "category": str(category).strip() if category else None,
            "priority": priority,
            "student_count_24": int(student_24) if student_24 and str(student_24).isdigit() else None,
            "student_count_25": int(student_25) if student_25 and str(student_25).isdigit() else None,
            "student_count_total": int(student_total) if student_total and str(student_total).isdigit() else None,
            "mentor_count": int(mentor_count) if mentor_count and str(mentor_count).isdigit() else None,
            "resident_leaders": resident_leaders,
            "degree_committee": degree_committee,
            "teaching_committee": teaching_committee,
            "mentors": mentors,
            "university_leaders": university_leaders,
            "notable_scholars": notable_scholars,
        }

    return universities_data


def merge_data() -> dict:
    """合并现有数据和 Excel 数据"""
    existing = load_existing_institutions()
    excel_data = parse_excel_data()

    print(f"📊 现有高校数: {len(existing.get('universities', []))}")
    print(f"📊 Excel 高校数: {len(excel_data)}")

    # 合并数据
    for uni in existing.get("universities", []):
        uni_name = uni.get("name", "")

        if uni_name in excel_data:
            # 保存原有的 org_name 和 departments
            original_org_name = uni.get("org_name")
            original_departments = uni.get("departments", [])

            # 合并 Excel 数据
            excel_info = excel_data[uni_name]
            uni.update(excel_info)

            # 恢复 org_name 和 departments
            uni["org_name"] = original_org_name
            uni["departments"] = original_departments

            print(f"  ✅ 合并: {uni_name}")
        else:
            print(f"  ⚠️  未找到 Excel 数据: {uni_name}")

    # 更新时间戳
    existing["last_updated"] = datetime.now(timezone.utc).isoformat()

    return existing


def main():
    """主函数"""
    print("🚀 开始构建完整的 institutions.json\n")

    # 检查文件
    if not INSTITUTIONS_FILE.exists():
        print(f"❌ 文件不存在: {INSTITUTIONS_FILE}")
        return

    if not EXCEL_FILE.exists():
        print(f"❌ 文件不存在: {EXCEL_FILE}")
        return

    # 合并数据
    merged_data = merge_data()

    # 保存
    with open(INSTITUTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(merged_data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 完成！文件已保存: {INSTITUTIONS_FILE}")

    # 统计
    total_unis = len(merged_data.get("universities", []))
    total_depts = sum(len(u.get("departments", [])) for u in merged_data.get("universities", []))
    total_scholars = sum(u.get("scholar_count", 0) for u in merged_data.get("universities", []))

    print(f"\n📊 统计:")
    print(f"  总高校数: {total_unis}")
    print(f"  总院系数: {total_depts}")
    print(f"  总学者数: {total_scholars}")


if __name__ == "__main__":
    main()
