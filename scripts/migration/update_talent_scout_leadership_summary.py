#!/usr/bin/env python3
# ruff: noqa: E501
from __future__ import annotations

import argparse
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TARGET = ROOT / "docs/files/人才画像数据汇总「4-27」.xlsx"
DEFAULT_SOURCES = ROOT / "sources/talent/talent_scout.yaml"

SUMMARY_SHEET = "01-执行摘要"
VIEW_SHEETS = {
    "02": "02-汇总明细",
    "03": "03-竞赛类视图",
    "04": "04-顶刊顶会类视图",
}
PERIOD_YEARS = ("2024", "2025")

GROUP_LABELS = {
    "talent_competitions": "竞赛",
    "talent_papers": "论文",
    "talent_github": "GitHub",
}

SOURCE_ID_PATTERN = re.compile(r"source_id=([A-Za-z0-9_\-]+)")

SOURCE_ALIASES = {
    "aliyun_tianchi": ["阿里天池"],
    "ccpc": ["CCPC 2024 Finals", "CCPC 中国大学生程序设计竞赛"],
    "casp": ["CASP13", "CASP14", "CASP15"],
    "igem": ["iGEM 2023", "iGEM 2024"],
    "noi_ioi": ["NOI/NOIP", "IOI 2024", "IOI 2025"],
    "academic_paper_authors": ["Academic Papers"],
}

SOURCE_PROFILES = {
    "ccf_bdci": ("部分覆盖", "已具备部分公告抓取能力", "DataFountain 详情页存在登录态/接口不稳定，当前以高校公告和公开名单补充。"),
    "aliyun_tianchi": ("部分覆盖", "已有历史/零散数据，暂未专项补爬", "榜单详情和登录态不稳定，本轮按要求未继续处理天池。"),
    "clue_benchmark": ("未覆盖", "暂未形成稳定自动爬虫", "榜单偏模型/团队，缺少可稳定落表的学生级姓名和学校字段。"),
    "semeval_conll": ("部分覆盖", "已具备部分公开页抓取能力", "页面多按论文、系统或团队发布，只有明确作者/成员姓名时入表。"),
    "neurips_competition": ("部分覆盖", "已具备部分公开页抓取能力", "年度页面结构不统一，部分记录只有项目或团队名。"),
    "robomaster": ("已覆盖", "已具备并落表", "已按公开名单/公告抽取学生级记录，后续继续补队伍成员页。"),
    "tencent_hok_ai": ("未覆盖", "暂未形成稳定自动爬虫", "公开完整成员名单不足，入口和登录态不稳定。"),
    "kaggle_grandmaster": ("已覆盖", "已具备并落表", "已能拉取榜单型信号；用户名到真实学生身份仍需交叉验证。"),
    "kdd_cup": ("部分覆盖", "已具备部分公开页抓取能力", "历年结果页格式差异较大，队伍成员字段不稳定。"),
    "icpc_rankings": ("已覆盖", "已具备并落表", "已按公开排名/名单抽取学生级竞赛记录。"),
    "huawei_elite_challenge": ("部分覆盖", "已具备高校公告抓取能力", "完整官方榜单入口和成员字段不稳定，当前以高校公告补充。"),
    "casp": ("部分覆盖", "已具备历史数据解析能力", "当前落表以历史赛季为主，24/25 学生级公开记录仍需补充。"),
    "igem": ("已覆盖", "已具备并落表", "已能从公开队伍/项目资料形成学生级或候选记录。"),
    "ecmwf_ai_challenge": ("未覆盖", "暂未形成稳定自动爬虫", "方向型信源较分散，公开页常只有团队或机构信息。"),
    "noi_ioi": ("已覆盖", "已具备并落表", "官网公告噪声较多，已保留非人名过滤后再入表。"),
    "ccpc": ("已覆盖", "已具备并落表", "已形成公开榜单/公告到学生记录的落表能力。"),
    "lanqiao": ("已覆盖", "已具备并落表", "已按归档文件解析，规模较大，需持续去重和噪声过滤。"),
    "asc_competition": ("部分覆盖", "已试抓，暂未落表", "公开页主要为队伍和学校层级，缺少稳定成员名单。"),
    "sc_student_cluster": ("已覆盖", "已具备并落表", "本轮新增学生级记录，后续扩展更多历史年度队伍页。"),
    "isc_student_cluster": ("已覆盖", "已具备并落表", "本轮新增学生级记录，公开页面分散，后续继续补高校报道。"),
    "pacman_huawei": ("未覆盖", "暂未形成稳定自动爬虫", "公开结果入口少，完整成员名单不稳定。"),
    "ctftime": ("未覆盖", "暂不作为主口径自动入表", "战队名、别名较多，真实姓名和学校映射弱；泛 CTF 仅作辅助线索。"),
    "ciscn_student": ("部分覆盖", "已具备部分公开页抓取能力", "CISCN 口径较稳定，但仍需避免泛 CTF 战队名直接进入姓名列。"),
    "os_kernel_competition": ("部分覆盖", "已具备部分公告抓取能力", "官方完整榜单和个人字段不统一，当前按高校公告和官方榜单双路径补齐。"),
    "icra_competition": ("部分覆盖", "已试抓，暂未落表", "公开页多为实验室、队伍或任务信息，姓名明确时再入表。"),
    "world_robot_contest": ("已覆盖", "已具备并落表", "已按公开名单抽取机器人竞赛学生级记录。"),
    "dblp_author": ("已覆盖", "已具备并落表", "已通过 DBLP/论文聚合数据形成作者候选记录。"),
    "arxiv_author": ("已覆盖", "已具备并落表", "已形成 arXiv 作者信号，作为论文侧辅助来源。"),
    "semantic_scholar_author": ("未覆盖", "已配置，未作为主口径", "存在限流和 API key 约束，本轮按要求不处理 Semantic Scholar。"),
    "openreview_author": ("已覆盖", "已具备并落表", "已覆盖 OpenReview 会议作者信号。"),
    "academic_paper_authors": ("已覆盖", "已具备并落表", "已接入本地论文聚合结果，作为顶刊顶会作者补充。"),
    "acl_anthology_author": ("已覆盖", "已具备并落表", "已直接解析 ACL Anthology 公开会议页。"),
    "cvf_openaccess_author": ("已覆盖", "已具备并落表", "已直接解析 CVF OpenAccess 公开会议页。"),
    "github_ai_users": ("已覆盖", "已具备并落表", "已形成 GitHub 用户候选信号；真实学生身份需交叉验证。"),
    "github_ai_repo_contributors": ("已覆盖", "已具备并落表", "已形成高星 AI 仓库贡献者候选信号；真实身份需交叉验证。"),
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def load_source_configs(path: Path) -> list[dict[str, Any]]:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    sources = config.get("sources") or []
    if not isinstance(sources, list):
        raise ValueError("sources/talent/talent_scout.yaml does not contain a source list")
    return sources


def is_period_row(row_text: str) -> bool:
    return any(year in row_text for year in PERIOD_YEARS)


def candidate_source_ids(row: tuple[Any, ...], source_ids: set[str]) -> set[str]:
    row_text = " ".join(clean_text(value) for value in row if value is not None)
    matched = {
        match.group(1)
        for match in SOURCE_ID_PATTERN.finditer(row_text)
        if match.group(1) in source_ids
    }
    if matched:
        return matched

    for source_id, aliases in SOURCE_ALIASES.items():
        if source_id in source_ids and any(alias in row_text for alias in aliases):
            matched.add(source_id)
    return matched


def count_view_rows(workbook: Any, sources: list[dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    source_ids = {source["id"] for source in sources}
    source_counts: dict[str, Any] = {
        source_id: {
            "all": Counter(),
            "period": Counter(),
        }
        for source_id in source_ids
    }
    view_totals: dict[str, int] = {}

    for short_name, sheet_name in VIEW_SHEETS.items():
        ws = workbook[sheet_name]
        total = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not any(clean_text(value) for value in row):
                continue
            total += 1
            row_text = " ".join(clean_text(value) for value in row if value is not None)
            for source_id in candidate_source_ids(row, source_ids):
                source_counts[source_id]["all"][short_name] += 1
                if is_period_row(row_text):
                    source_counts[source_id]["period"][short_name] += 1
        view_totals[short_name] = total

    return source_counts, view_totals


def format_counts(counts: Counter[str]) -> str:
    if not counts:
        return "0"
    tokens = [f"{sheet}:{counts.get(sheet, 0)}" for sheet in VIEW_SHEETS if counts.get(sheet, 0)]
    return "；".join(tokens) if tokens else "0"


def source_rows(sources: list[dict[str, Any]], source_counts: dict[str, Any]) -> list[list[Any]]:
    rows = []
    for index, source in enumerate(sources, start=1):
        source_id = source["id"]
        status, capability, note = SOURCE_PROFILES[source_id]
        category = GROUP_LABELS.get(source.get("group"), clean_text(source.get("group")))
        all_count = format_counts(source_counts[source_id]["all"])
        period_count = format_counts(source_counts[source_id]["period"])
        rows.append(
            [
                index,
                source_id,
                source.get("name"),
                category,
                status,
                capability,
                f"全部 {all_count}；24/25 {period_count}",
                note,
            ]
        )
    return rows


def write_row(ws: Any, row_idx: int, values: list[Any]) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row_idx, col_idx).value = value


def clear_summary(ws: Any) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_summary(ws: Any, max_row: int) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="BDD7EE")

    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    for col in range(1, 9):
        ws.cell(1, col).fill = title_fill

    for row in (5, 8, 18):
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", bold=True)
            cell.fill = section_fill

    for row in (9, 19):
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", bold=True)
            cell.fill = header_fill

    widths = [8, 27, 30, 10, 12, 24, 28, 54]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 30 if row >= 19 else 24
        for col in range(1, 9):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A19"
    ws.auto_filter.ref = f"A19:H{max_row}"


def update_summary(workbook: Any, sources: list[dict[str, Any]]) -> dict[str, Any]:
    ws = workbook[SUMMARY_SHEET]
    source_counts, view_totals = count_view_rows(workbook, sources)
    rows = source_rows(sources, source_counts)

    status_counts = Counter(row[4] for row in rows)
    period_source_count = sum(
        1 for source_id in source_counts if sum(source_counts[source_id]["period"].values()) > 0
    )
    sc_isc_added = (
        source_counts["sc_student_cluster"]["all"].get("02", 0)
        + source_counts["isc_student_cluster"]["all"].get("02", 0)
    )

    clear_summary(ws)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    write_row(ws, 1, ["人才画像数据汇总执行摘要"])
    write_row(ws, 2, ["面向领导汇报：按具体信源逐行列示爬虫能力、落表情况和未覆盖原因。"])
    write_row(ws, 3, ["更新时间", generated_at])

    write_row(ws, 5, ["一、汇报结论"])
    conclusion = (
        f"本版摘要已按系统配置的 {len(sources)} 个具体信源逐行列示，不再按需求域合并。"
        f"当前已覆盖 {status_counts['已覆盖']} 个、部分覆盖 {status_counts['部分覆盖']} 个、"
        f"未覆盖 {status_counts['未覆盖']} 个；其中 {period_source_count} 个信源已有 2024/2025 年度落表记录。"
        f"SC/ISC 超算赛本轮新增 {sc_isc_added} 条学生级记录。未覆盖项主要受限于公开页缺少学生级姓名、"
        "仅提供团队/学校信息、需要登录态/API key 或反爬限制。"
    )
    write_row(ws, 6, ["结论", conclusion])

    write_row(ws, 8, ["二、核心量化指标"])
    write_row(ws, 9, ["指标", "数值", "口径说明", "状态"])
    metrics = [
        ["汇总明细有效记录数", view_totals["02"], "02-汇总明细", "已统计"],
        ["竞赛类视图有效记录数", view_totals["03"], "03-竞赛类视图", "已统计"],
        ["顶刊顶会类视图有效记录数", view_totals["04"], "04-顶刊顶会类视图", "已统计"],
        ["具体信源配置数", len(sources), "sources/talent/talent_scout.yaml", "已统计"],
        ["已覆盖 / 部分覆盖 / 未覆盖", f"{status_counts['已覆盖']} / {status_counts['部分覆盖']} / {status_counts['未覆盖']}", "逐 source_id 统计", "已统计"],
        ["有 2024/2025 落表记录的信源数", period_source_count, "按各视图行文本中的 2024/2025 判断", "已统计"],
        ["本轮新增 SC/ISC 学生级记录", sc_isc_added, "已写入 02/03 视图", "已处理"],
    ]
    for offset, metric in enumerate(metrics, start=10):
        write_row(ws, offset, metric)

    write_row(ws, 18, ["三、逐信源爬虫能力明细"])
    write_row(ws, 19, ["序号", "source_id", "信源名称", "类别", "覆盖状态", "爬虫能力", "落表记录", "说明"])
    for offset, row in enumerate(rows, start=20):
        write_row(ws, offset, row)

    max_row = 19 + len(rows)
    style_summary(ws, max_row)
    return {
        "source_count": len(sources),
        "status_counts": dict(status_counts),
        "period_source_count": period_source_count,
        "view_totals": view_totals,
        "max_row": max_row,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Update talent scout leadership summary sheet.")
    parser.add_argument("--target", default=str(DEFAULT_TARGET), help="Target workbook path")
    parser.add_argument("--sources", default=str(DEFAULT_SOURCES), help="talent_scout.yaml path")
    args = parser.parse_args()

    target_path = Path(args.target).resolve()
    sources_path = Path(args.sources).resolve()
    sources = load_source_configs(sources_path)
    missing_profiles = sorted({source["id"] for source in sources} - set(SOURCE_PROFILES))
    if missing_profiles:
        raise ValueError(f"Missing source profiles: {', '.join(missing_profiles)}")

    workbook = load_workbook(target_path)
    result = update_summary(workbook, sources)
    workbook.save(target_path)
    print(result)


if __name__ == "__main__":
    main()
