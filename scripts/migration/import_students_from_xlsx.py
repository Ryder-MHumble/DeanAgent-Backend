#!/usr/bin/env python3
"""Import supervised students from xlsx files into supervised_students table.

Priority order:
1) 24级学生与导师信息.xlsx
2) 25级学生与导师信息.xlsx
3) 26级学生与导师信息.xlsx
4) all_student.xlsx (supplement only)

Usage:
  ./.venv/bin/python scripts/migration/import_students_from_xlsx.py --dry-run
  ./.venv/bin/python scripts/migration/import_students_from_xlsx.py --apply
"""

from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from app.config import settings
from app.db.pool import close_pool, get_pool, init_pool

BASE_DIR = Path("data/student")
FILE_24 = BASE_DIR / "24级学生与导师信息.xlsx"
FILE_25 = BASE_DIR / "25级学生与导师信息.xlsx"
FILE_26 = BASE_DIR / "26级学生与导师信息.xlsx"
FILE_ALL = BASE_DIR / "all_student.xlsx"
UNKNOWN_MENTOR_SCHOLAR_ID = "__unknown_student_mentor__"
UNKNOWN_MENTOR_SCHOLAR_NAME = "待匹配导师"


def n(value: Any) -> str:
    return "".join(str(value or "").replace("\u3000", " ").strip().split()).lower()


def clean(value: Any) -> str:
    return str(value or "").strip()


def year_from_text(value: Any) -> int | None:
    token = clean(value)
    if not token:
        return None
    for part in [token, token[:4]]:
        try:
            year = int(float(part))
        except Exception:
            continue
        if 1900 <= year <= 2100:
            return year
    digits = "".join(ch for ch in token if ch.isdigit())
    if len(digits) >= 4:
        year = int(digits[:4])
        if 1900 <= year <= 2100:
            return year
    return None


def load_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean(cell) for cell in header_row]
    rows: list[dict[str, str]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row: dict[str, str] = {}
        for idx, cell in enumerate(r):
            if idx >= len(headers):
                continue
            key = headers[idx]
            if key:
                row[key] = clean(cell)
        rows.append(row)
    return rows


@dataclass
class StudentCandidate:
    source: str
    source_priority: int
    student_no: str
    name: str
    enrollment_year: int | None
    degree_type: str
    home_university: str
    major: str
    email: str
    phone: str
    mentor_name: str
    mentor_email: str
    mentor_title: str
    notes: str

    def key_no(self) -> str:
        return clean(self.student_no)

    def key_name_year_school(self) -> str:
        return f"{n(self.name)}|{self.enrollment_year or ''}|{n(self.home_university)}"

    def key_name_year(self) -> str:
        return f"{n(self.name)}|{self.enrollment_year or ''}"


def row_24_to_candidate(row: dict[str, str]) -> StudentCandidate | None:
    name = clean(row.get("学生姓名"))
    if not name:
        return None
    mentor_name = clean(row.get("导师姓名"))
    mentor_title = clean(row.get("导师职称"))
    mentor_email = clean(row.get("导师邮箱"))
    return StudentCandidate(
        source=FILE_24.name,
        source_priority=1,
        student_no=clean(row.get("学号")),
        name=name,
        enrollment_year=2024,
        degree_type="",
        home_university=clean(row.get("推荐高校")),
        major="",
        email="",
        phone="",
        mentor_name=mentor_name,
        mentor_email=mentor_email,
        mentor_title=mentor_title,
        notes=f"source={FILE_24.name}",
    )


def row_25_to_candidate(row: dict[str, str]) -> StudentCandidate | None:
    name = clean(row.get("姓名"))
    if not name:
        return None
    mentor_name = clean(row.get("拟报博士导师姓名"))
    mentor_title = clean(row.get("导师职称"))
    mentor_email = clean(row.get("导师邮箱"))
    return StudentCandidate(
        source=FILE_25.name,
        source_priority=2,
        student_no=clean(row.get("学号")),
        name=name,
        enrollment_year=2025,
        degree_type="",
        home_university=clean(row.get("学籍高校")),
        major="",
        email="",
        phone="",
        mentor_name=mentor_name,
        mentor_email=mentor_email,
        mentor_title=mentor_title,
        notes=f"source={FILE_25.name}",
    )


def row_26_to_candidate(row: dict[str, str]) -> StudentCandidate | None:
    name = clean(row.get("姓名"))
    if not name:
        return None
    degree_type = clean(row.get("学生类别"))
    return StudentCandidate(
        source=FILE_26.name,
        source_priority=3,
        student_no="",
        name=name,
        enrollment_year=year_from_text(degree_type) or 2026,
        degree_type=degree_type,
        home_university=clean(row.get("确认接受高校")),
        major=clean(row.get("确认接受专业")),
        email="",
        phone="",
        mentor_name=clean(row.get("导师姓名")),
        mentor_email="",
        mentor_title="",
        notes=f"source={FILE_26.name};batch={clean(row.get('批次'))}",
    )


def row_all_to_candidate(row: dict[str, str]) -> StudentCandidate | None:
    name = clean(row.get("学生姓名"))
    if not name:
        return None
    mentor_name = clean(row.get("高校导师")) or clean(row.get("学院导师"))
    return StudentCandidate(
        source=FILE_ALL.name,
        source_priority=4,
        student_no="",
        name=name,
        enrollment_year=year_from_text(row.get("年级")),
        degree_type=clean(row.get("培养类型")),
        home_university=clean(row.get("学籍学校")),
        major=clean(row.get("专业")),
        email=clean(row.get("邮箱")),
        phone=clean(row.get("手机")),
        mentor_name=mentor_name,
        mentor_email="",
        mentor_title="",
        notes=f"source={FILE_ALL.name}",
    )


def merge_candidate(base: StudentCandidate, incoming: StudentCandidate) -> StudentCandidate:
    def pick(a: str, b: str) -> str:
        return a if clean(a) else clean(b)

    base.student_no = pick(base.student_no, incoming.student_no)
    base.home_university = pick(base.home_university, incoming.home_university)
    base.degree_type = pick(base.degree_type, incoming.degree_type)
    base.major = pick(base.major, incoming.major)
    base.email = pick(base.email, incoming.email)
    base.phone = pick(base.phone, incoming.phone)
    base.mentor_name = pick(base.mentor_name, incoming.mentor_name)
    base.mentor_email = pick(base.mentor_email, incoming.mentor_email)
    base.mentor_title = pick(base.mentor_title, incoming.mentor_title)
    if base.enrollment_year is None and incoming.enrollment_year is not None:
        base.enrollment_year = incoming.enrollment_year
    if incoming.source_priority < base.source_priority:
        base.source_priority = incoming.source_priority
        base.source = incoming.source
    if incoming.notes and incoming.notes not in base.notes:
        base.notes = f"{base.notes};{incoming.notes}" if base.notes else incoming.notes
    return base


def build_candidates() -> tuple[list[StudentCandidate], dict[str, int]]:
    stats = {
        "raw_24": 0,
        "raw_25": 0,
        "raw_26": 0,
        "raw_all": 0,
        "merged_updates": 0,
        "new_records": 0,
    }
    selected: list[StudentCandidate] = []
    by_no: dict[str, int] = {}
    by_nys: dict[str, int] = {}
    by_ny: dict[str, int] = {}

    def add_or_merge(c: StudentCandidate) -> None:
        key_no = c.key_no()
        key_nys = c.key_name_year_school()
        key_ny = c.key_name_year()

        hit: int | None = None
        if key_no and key_no in by_no:
            hit = by_no[key_no]
        elif key_nys in by_nys:
            hit = by_nys[key_nys]
        elif key_ny in by_ny:
            hit = by_ny[key_ny]

        if hit is not None:
            selected[hit] = merge_candidate(selected[hit], c)
            stats["merged_updates"] += 1
            base = selected[hit]
            if base.key_no():
                by_no[base.key_no()] = hit
            by_nys[base.key_name_year_school()] = hit
            by_ny[base.key_name_year()] = hit
            return

        idx = len(selected)
        selected.append(c)
        stats["new_records"] += 1
        if key_no:
            by_no[key_no] = idx
        by_nys[key_nys] = idx
        by_ny[key_ny] = idx

    for row in load_rows(FILE_24):
        stats["raw_24"] += 1
        c = row_24_to_candidate(row)
        if c is not None:
            add_or_merge(c)
    for row in load_rows(FILE_25):
        stats["raw_25"] += 1
        c = row_25_to_candidate(row)
        if c is not None:
            add_or_merge(c)
    for row in load_rows(FILE_26):
        stats["raw_26"] += 1
        c = row_26_to_candidate(row)
        if c is not None:
            add_or_merge(c)
    for row in load_rows(FILE_ALL):
        stats["raw_all"] += 1
        c = row_all_to_candidate(row)
        if c is not None:
            add_or_merge(c)

    return selected, stats


@dataclass
class ScholarLite:
    scholar_id: str
    name: str
    university: str
    email: str
    is_school_mentor: bool


def pick_mentor(
    candidate: StudentCandidate,
    scholar_by_name: dict[str, list[ScholarLite]],
) -> tuple[str | None, str]:
    raw_mentor_name = clean(candidate.mentor_name)
    if not raw_mentor_name:
        return None, "mentor_name_empty"

    placeholder_tokens = {"未确定", "未确认", "待定", "待确认"}
    normalized = raw_mentor_name
    for sep in ("／", "/", "+", "、", "，", ",", "；", ";", "|"):
        normalized = normalized.replace(sep, "|")
    tokens = [clean(t) for t in normalized.split("|") if clean(t)]
    mentor_name_candidates: list[str] = []
    for token in tokens or [raw_mentor_name]:
        if token in placeholder_tokens:
            continue
        if "指导小组" in token:
            continue
        mentor_name_candidates.append(token)
    if not mentor_name_candidates:
        return None, "mentor_name_empty"

    mentor_email = n(candidate.mentor_email)
    home_uni = n(candidate.home_university)

    for mentor_name in mentor_name_candidates:
        options = scholar_by_name.get(n(mentor_name), [])
        if not options:
            continue

        if mentor_email:
            email_hits = [s for s in options if n(s.email) == mentor_email]
            if len(email_hits) == 1:
                return email_hits[0].scholar_id, "email_match"
            if len(email_hits) > 1:
                options = email_hits

        if home_uni:
            uni_hits = [
                s
                for s in options
                if home_uni and (home_uni in n(s.university) or n(s.university) in home_uni)
            ]
            if len(uni_hits) == 1:
                return uni_hits[0].scholar_id, "university_match"
            if len(uni_hits) > 1:
                options = uni_hits

        if len(options) == 1:
            return options[0].scholar_id, "single_name_hit"

        ranked = sorted(
            options,
            key=lambda s: (
                1 if s.is_school_mentor else 0,
                1 if clean(s.email) else 0,
                len(n(s.university)),
                s.scholar_id,
            ),
            reverse=True,
        )
        if not ranked:
            continue
        if len(ranked) >= 2:
            top = ranked[0]
            second = ranked[1]
            top_score = (
                1 if top.is_school_mentor else 0,
                1 if clean(top.email) else 0,
                len(n(top.university)),
            )
            sec_score = (
                1 if second.is_school_mentor else 0,
                1 if clean(second.email) else 0,
                len(n(second.university)),
            )
            if top_score == sec_score:
                # If duplicate rows share same name/university, pick deterministic first id.
                if n(top.name) == n(second.name) and n(top.university) == n(second.university):
                    best = sorted(ranked, key=lambda s: s.scholar_id)[0]
                    return best.scholar_id, "duplicate_name_university_pick"
                continue
        return ranked[0].scholar_id, "ranked_pick"

    if any(t in placeholder_tokens for t in tokens):
        return None, "mentor_name_empty"
    return None, "mentor_not_found"


async def connect_pool() -> None:
    if settings.POSTGRES_DSN:
        await init_pool(dsn=settings.POSTGRES_DSN)
    else:
        await init_pool(
            host=settings.POSTGRES_HOST,
            port=settings.POSTGRES_PORT,
            user=settings.POSTGRES_USER,
            password=settings.POSTGRES_PASSWORD,
            database=settings.POSTGRES_DB,
        )


async def run(apply_changes: bool) -> None:
    candidates, local_stats = build_candidates()
    await connect_pool()
    pool = get_pool()

    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO scholars (id, name)
            VALUES ($1, $2)
            ON CONFLICT (id) DO UPDATE
            SET name = EXCLUDED.name,
                updated_at = now()
            """,
            UNKNOWN_MENTOR_SCHOLAR_ID,
            UNKNOWN_MENTOR_SCHOLAR_NAME,
        )

        scholar_rows = await conn.fetch(
            """
            SELECT id, name, university, email, project_category, project_subcategory
            FROM scholars
            """
        )
        scholar_by_name: dict[str, list[ScholarLite]] = {}
        for r in scholar_rows:
            s = ScholarLite(
                scholar_id=str(r["id"]),
                name=clean(r["name"]),
                university=clean(r["university"]),
                email=clean(r["email"]),
                is_school_mentor=(
                    clean(r["project_category"]) == "教育培养"
                    and clean(r["project_subcategory"]) == "学院学生高校导师"
                ),
            )
            if s.name:
                scholar_by_name.setdefault(n(s.name), []).append(s)

        existing_rows = await conn.fetch(
            """
            SELECT id, scholar_id, student_no, name, enrollment_year, home_university
            FROM supervised_students
            """
        )
        existing_by_no: dict[str, str] = {}
        existing_by_nyu: dict[str, str] = {}
        for r in existing_rows:
            row_id = clean(r["id"])
            student_no = clean(r["student_no"])
            if student_no:
                existing_by_no[student_no] = row_id
            key = f"{n(r['name'])}|{r['enrollment_year'] or ''}|{n(r['home_university'])}"
            existing_by_nyu[key] = row_id

        mapped = 0
        unresolved = 0
        inserted = 0
        updated = 0
        unresolved_samples: list[str] = []
        reason_counter: dict[str, int] = {}

        for c in candidates:
            if not clean(c.name):
                continue
            scholar_id, reason = pick_mentor(c, scholar_by_name)
            if scholar_id is None:
                unresolved += 1
                reason_counter[reason] = reason_counter.get(reason, 0) + 1
                if len(unresolved_samples) < 30:
                    unresolved_line = (
                        f"{c.name} | year={c.enrollment_year or ''} | "
                        f"mentor={c.mentor_name or '(空)'} | "
                        f"source={c.source} | reason={reason}"
                    )
                    unresolved_samples.append(
                        unresolved_line
                    )
                scholar_id = UNKNOWN_MENTOR_SCHOLAR_ID
            else:
                mapped += 1

            note_parts = [clean(c.notes)]
            if c.mentor_title:
                note_parts.append(f"mentor_title={c.mentor_title}")
            if c.mentor_email:
                note_parts.append(f"mentor_email={c.mentor_email}")
            notes = ";".join([p for p in note_parts if p])

            payload = {
                "scholar_id": scholar_id,
                "student_no": clean(c.student_no) or None,
                "name": clean(c.name),
                "home_university": clean(c.home_university) or None,
                "major": clean(c.major) or None,
                "degree_type": clean(c.degree_type) or None,
                "enrollment_year": c.enrollment_year,
                "expected_graduation_year": None,
                "status": "在读",
                "email": clean(c.email) or None,
                "phone": clean(c.phone) or None,
                "notes": notes or None,
                "mentor_name": clean(c.mentor_name) or None,
                "added_by": "system:student_import",
            }

            existing_id: str | None = None
            if payload["student_no"] and payload["student_no"] in existing_by_no:
                existing_id = existing_by_no[payload["student_no"]]
            if existing_id is None:
                key = (
                    f"{n(payload['name'])}|"
                    f"{payload['enrollment_year'] or ''}|"
                    f"{n(payload['home_university'])}"
                )
                existing_id = existing_by_nyu.get(key)

            if not apply_changes:
                if existing_id:
                    updated += 1
                else:
                    inserted += 1
                continue

            if existing_id:
                await conn.execute(
                    """
                    UPDATE supervised_students
                    SET
                      scholar_id = $2,
                      student_no = $3,
                      name = $4,
                      home_university = $5,
                      major = $6,
                      degree_type = $7,
                      enrollment_year = $8,
                      expected_graduation_year = $9,
                      status = $10,
                      email = $11,
                      phone = $12,
                      notes = $13,
                      mentor_name = $14,
                      added_by = $15,
                      updated_at = now()
                    WHERE id = $1
                    """,
                    existing_id,
                    payload["scholar_id"],
                    payload["student_no"],
                    payload["name"],
                    payload["home_university"],
                    payload["major"],
                    payload["degree_type"],
                    payload["enrollment_year"],
                    payload["expected_graduation_year"],
                    payload["status"],
                    payload["email"],
                    payload["phone"],
                    payload["notes"],
                    payload["mentor_name"],
                    payload["added_by"],
                )
                updated += 1
            else:
                row = await conn.fetchrow(
                    """
                    INSERT INTO supervised_students
                    (scholar_id, student_no, name, home_university, major, degree_type,
                     enrollment_year, expected_graduation_year, status, email, phone, notes,
                     mentor_name, added_by)
                    VALUES
                    ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14)
                    RETURNING id
                    """,
                    payload["scholar_id"],
                    payload["student_no"],
                    payload["name"],
                    payload["home_university"],
                    payload["major"],
                    payload["degree_type"],
                    payload["enrollment_year"],
                    payload["expected_graduation_year"],
                    payload["status"],
                    payload["email"],
                    payload["phone"],
                    payload["notes"],
                    payload["mentor_name"],
                    payload["added_by"],
                )
                new_id = clean(row["id"])
                inserted += 1
                if payload["student_no"]:
                    existing_by_no[payload["student_no"]] = new_id
                key = (
                    f"{n(payload['name'])}|"
                    f"{payload['enrollment_year'] or ''}|"
                    f"{n(payload['home_university'])}"
                )
                existing_by_nyu[key] = new_id

        total_students = await conn.fetchval("SELECT COUNT(*)::bigint FROM supervised_students")

    print("")
    print("===== Student Import =====")
    print(f"Mode: {'APPLY' if apply_changes else 'DRY-RUN'}")
    print(f"source_rows_24: {local_stats['raw_24']}")
    print(f"source_rows_25: {local_stats['raw_25']}")
    print(f"source_rows_26: {local_stats['raw_26']}")
    print(f"source_rows_all: {local_stats['raw_all']}")
    print(f"candidate_new_records: {local_stats['new_records']}")
    print(f"candidate_merged_updates: {local_stats['merged_updates']}")
    print(f"mentor_mapped_records: {mapped}")
    print(f"mentor_unresolved_records: {unresolved}")
    print(f"rows_inserted: {inserted}")
    print(f"rows_updated: {updated}")
    print(f"supervised_students_table_count: {total_students}")
    if reason_counter:
        print("unresolved_reason_breakdown:")
        for reason, cnt in sorted(reason_counter.items(), key=lambda x: (-x[1], x[0])):
            print(f"  - {reason}: {cnt}")
    if unresolved_samples:
        print("unresolved_samples:")
        for line in unresolved_samples:
            print(f"  - {line}")

    await close_pool()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import student xlsx files into supervised_students"
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Preview only")
    mode.add_argument("--apply", action="store_true", help="Apply import")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    asyncio.run(run(apply_changes=bool(args.apply)))


if __name__ == "__main__":
    main()
