#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
EXPORT_DIR = ROOT / "data" / "exports"
DEFAULT_TARGET = EXPORT_DIR / "人才画像数据汇总（正式汇报版）.xlsx"
STUDENT_SHEET = "02-汇总明细"
CANDIDATE_SHEET = "03-人才候选池明细"
MAPPING_SHEET = "04-字段映射"
LEGACY_MAPPING_SHEET = "03-字段映射"
SOURCE_SHEET = "TalentSignals"
SOURCE_PREFIX = "crawl_results_"

STUDENT_HEADERS = [
    "学生姓名",
    "高校",
    "院系",
    "邮箱",
    "赛事名称",
    "赛事全称",
    "奖项等级",
    "比赛时间",
    "赛题方向",
    "来源URL",
    "数据来源",
    "备注",
    "来源文件",
    "来源Sheet",
    "来源类型",
]

CANDIDATE_HEADERS = [
    "候选人姓名",
    "学校",
    "院系",
    "邮箱",
    "信号名称",
    "信号全称",
    "级别/排序",
    "时间",
    "方向",
    "来源URL",
    "来源名称",
    "备注",
    "来源文件",
    "来源Sheet",
    "来源类型",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def optional_text(value: Any) -> str | None:
    token = clean_text(value)
    return token or None


def latest_crawl_workbook(export_dir: Path) -> Path:
    matches = sorted(export_dir.glob(f"{SOURCE_PREFIX}*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No {SOURCE_PREFIX}*.xlsx found under data/exports")
    return matches[-1]


def is_crawl_result_row(row: list[Any] | tuple[Any, ...]) -> bool:
    return clean_text(row[12]).startswith(SOURCE_PREFIX) and clean_text(row[13]) == SOURCE_SHEET


def copy_row_style(ws, template_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(template_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def copy_column_dimensions(source_ws, target_ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = chr(64 + col)
        source_dim = source_ws.column_dimensions[letter]
        target_dim = target_ws.column_dimensions[letter]
        target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden


def infer_signal_type(row: dict[str, Any]) -> str:
    if clean_text(row.get("competition_name")):
        return "competition"
    if clean_text(row.get("paper_title")) or clean_text(row.get("venue")):
        return "paper_author"
    if clean_text(row.get("github_login")) or clean_text(row.get("repo_full_name")):
        return "github_contributor"
    return "generic"


def build_signal_name(row: dict[str, Any], signal_type: str) -> str:
    if signal_type == "competition":
        return clean_text(row.get("competition_name")) or clean_text(row.get("source_name"))
    if signal_type == "paper_author":
        return clean_text(row.get("venue")) or clean_text(row.get("source_name")) or "论文作者信号"
    if signal_type == "github_contributor":
        return clean_text(row.get("repo_full_name")) or clean_text(row.get("source_name")) or "GitHub 候选池"
    return clean_text(row.get("evidence_title")) or clean_text(row.get("source_name"))


def build_signal_full_name(row: dict[str, Any], signal_type: str, signal_name: str) -> str:
    if signal_type == "competition":
        return clean_text(row.get("evidence_title")) or signal_name
    if signal_type == "paper_author":
        return (
            clean_text(row.get("paper_title"))
            or clean_text(row.get("evidence_title"))
            or signal_name
        )
    if signal_type == "github_contributor":
        return (
            clean_text(row.get("repo_full_name"))
            or clean_text(row.get("evidence_title"))
            or signal_name
        )
    return clean_text(row.get("evidence_title")) or signal_name


def build_level_or_rank(row: dict[str, Any], signal_type: str) -> str | None:
    award_level = clean_text(row.get("award_level"))
    ranking = clean_text(row.get("ranking"))
    author_order = clean_text(row.get("author_order"))
    contributions = clean_text(row.get("contributions"))
    followers = clean_text(row.get("followers"))
    citation_count = clean_text(row.get("citation_count"))

    if signal_type == "competition":
        if award_level and ranking:
            return f"{award_level} / 排名 {ranking}"
        if award_level:
            return award_level
        if ranking:
            return f"排名 {ranking}"
        return None
    if signal_type == "paper_author":
        tokens: list[str] = []
        if author_order:
            tokens.append(f"作者序 {author_order}")
        if citation_count:
            tokens.append(f"引用 {citation_count}")
        return " / ".join(tokens) or None
    if signal_type == "github_contributor":
        tokens = []
        if contributions:
            tokens.append(f"贡献 {contributions}")
        if followers:
            tokens.append(f"关注者 {followers}")
        return " / ".join(tokens) or None
    return None


def build_time_token(row: dict[str, Any]) -> str | None:
    for field in ("season_year", "venue_year"):
        token = clean_text(row.get(field))
        if token:
            return token
    executed_at = clean_text(row.get("executed_at"))
    if not executed_at:
        return None
    try:
        return datetime.fromisoformat(executed_at.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return executed_at[:10]


def build_candidate_note(row: dict[str, Any], signal_type: str, signal_full_name: str) -> str | None:
    pairs: list[tuple[str, str | None]] = [
        ("信号类型", signal_type),
        ("source_id", optional_text(row.get("source_id"))),
        ("record_status", optional_text(row.get("record_status"))),
        ("confidence", optional_text(row.get("confidence"))),
        ("track", optional_text(row.get("track"))),
        ("team_name", optional_text(row.get("team_name"))),
        ("paper_count_in_scope", optional_text(row.get("paper_count_in_scope"))),
        ("dblp_pid", optional_text(row.get("dblp_pid"))),
        ("github_login", optional_text(row.get("github_login"))),
        ("company", optional_text(row.get("company"))),
        ("blog", optional_text(row.get("blog"))),
        ("executed_at", optional_text(row.get("executed_at"))),
    ]
    evidence_title = optional_text(row.get("evidence_title"))
    if evidence_title and evidence_title != signal_full_name:
        pairs.append(("evidence_title", evidence_title))
    tokens = [f"{key}={value}" for key, value in pairs if value]
    return "；".join(tokens) if tokens else None


def build_candidate_row(row: dict[str, Any], source_file_name: str) -> list[Any]:
    signal_type = infer_signal_type(row)
    signal_name = build_signal_name(row, signal_type)
    signal_full_name = build_signal_full_name(row, signal_type, signal_name)
    return [
        optional_text(row.get("candidate_name")) or optional_text(row.get("github_login")),
        optional_text(row.get("university")),
        optional_text(row.get("department")),
        optional_text(row.get("email")),
        signal_name or None,
        signal_full_name or None,
        build_level_or_rank(row, signal_type),
        build_time_token(row),
        optional_text(row.get("track")),
        optional_text(row.get("evidence_url")),
        optional_text(row.get("source_name")),
        build_candidate_note(row, signal_type, signal_full_name),
        source_file_name,
        SOURCE_SHEET,
        "Excel",
    ]


def candidate_row_key(values: list[Any]) -> tuple[str, ...]:
    return tuple(clean_text(value).lower() for value in values[:11])


def collect_candidate_rows(source_path: Path) -> tuple[list[list[Any]], dict[str, int]]:
    workbook = load_workbook(source_path, data_only=True)
    ws = workbook[SOURCE_SHEET]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows: list[list[Any]] = []
    seen_keys: set[tuple[str, ...]] = set()
    stats = {
        "source_total_rows": max(ws.max_row - 1, 0),
        "structured_rows": 0,
        "skipped_non_structured": 0,
        "skipped_missing_name": 0,
        "source_deduped": 0,
    }

    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values, strict=False))
        if clean_text(row.get("record_status")) != "structured":
            stats["skipped_non_structured"] += 1
            continue
        if not clean_text(row.get("candidate_name")) and not clean_text(row.get("github_login")):
            stats["skipped_missing_name"] += 1
            continue
        candidate_row = build_candidate_row(row, source_path.name)
        key = candidate_row_key(candidate_row)
        if key in seen_keys:
            stats["source_deduped"] += 1
            continue
        seen_keys.add(key)
        rows.append(candidate_row)
        stats["structured_rows"] += 1

    return rows, stats


def read_detail_rows(ws) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=5, max_col=len(STUDENT_HEADERS), values_only=True):
        if any(clean_text(value) for value in row):
            rows.append(list(row))
    return rows


def clear_data_rows(ws, max_col: int) -> None:
    for row in range(5, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_rows(ws, rows: list[list[Any]], max_col: int) -> None:
    clear_data_rows(ws, max_col)
    for row_idx, values in enumerate(rows, start=5):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value
    last_row = max(4, len(rows) + 4)
    ws.auto_filter.ref = f"A4:{chr(64 + max_col)}{last_row}"


def rebuild_table(ws, *, table_name: str, headers: list[str], max_col: int, row_count: int) -> None:
    for existing_name in list(ws.tables.keys()):
        del ws.tables[existing_name]
    last_row = max(5, row_count + 4)
    ref = f"A4:{chr(64 + max_col)}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = ref
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col).value = header


def resize_table(ws, table_name: str, ref: str) -> None:
    if table_name in ws.tables:
        ws.tables[table_name].ref = ref


def ensure_candidate_sheet(workbook, student_ws):
    if MAPPING_SHEET not in workbook.sheetnames and LEGACY_MAPPING_SHEET in workbook.sheetnames:
        workbook[LEGACY_MAPPING_SHEET].title = MAPPING_SHEET

    if CANDIDATE_SHEET in workbook.sheetnames:
        candidate_ws = workbook[CANDIDATE_SHEET]
    else:
        candidate_ws = workbook.copy_worksheet(student_ws)
        candidate_ws.title = CANDIDATE_SHEET

    copy_column_dimensions(student_ws, candidate_ws, len(CANDIDATE_HEADERS))
    candidate_ws.freeze_panes = student_ws.freeze_panes
    candidate_ws.sheet_view.showGridLines = student_ws.sheet_view.showGridLines
    candidate_ws["A1"] = "人才候选池明细"
    candidate_ws["A2"] = "说明：本页收录 crawl_results 导出的潜在引入人才信号，按竞赛、论文、GitHub 等外部信源统一汇总。"
    for col, header in enumerate(CANDIDATE_HEADERS, start=1):
        candidate_ws.cell(4, col).value = header

    if MAPPING_SHEET in workbook.sheetnames:
        mapping_ws = workbook[MAPPING_SHEET]
        sheets = workbook._sheets
        candidate_idx = sheets.index(candidate_ws)
        mapping_idx = sheets.index(mapping_ws)
        if candidate_idx > mapping_idx:
            sheets.insert(mapping_idx, sheets.pop(candidate_idx))

    return candidate_ws


def compute_source_line_stats(summary_ws) -> tuple[int, int]:
    excel_rows = 0
    json_rows = 0
    for row in range(6, 18):
        source_name = clean_text(summary_ws.cell(row, 5).value)
        count = summary_ws.cell(row, 7).value
        if not source_name:
            continue
        count_value = int(count or 0)
        if source_name.endswith(".json") or ".json:" in source_name:
            json_rows += count_value
        elif source_name.endswith(".xlsx:TalentSignals"):
            continue
        else:
            excel_rows += count_value
    return excel_rows, json_rows


def update_summary_sheet(
    summary_ws,
    *,
    source_path: Path,
    candidate_count: int,
    student_total: int,
    source_total_rows: int,
) -> None:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    source_name = f"{source_path.name}:{SOURCE_SHEET}"

    source_previously_listed = False
    source_row = None
    for row in range(6, 18):
        if clean_text(summary_ws.cell(row, 5).value) == source_name:
            source_row = row
            source_previously_listed = True
            break
    if source_row is None:
        source_row = 10

    summary_ws.cell(source_row, 5).value = source_name
    summary_ws.cell(source_row, 6).value = source_total_rows
    summary_ws.cell(source_row, 7).value = candidate_count
    summary_ws.cell(source_row, 8).value = "纳入人才候选池视图，不计入学生画像视图汇总总量。"

    excel_rows, json_rows = compute_source_line_stats(summary_ws)
    summary_ws["A9"] = "Excel纳入学生视图记录数"
    summary_ws["A11"] = "学生画像视图总记录数"
    copy_row_style(summary_ws, 11, 16, 10)
    copy_row_style(summary_ws, 12, 17, 10)
    summary_ws["A16"] = "人才候选池视图记录数"
    summary_ws["A17"] = "双视图总记录数"

    summary_ws["B6"] = generated_at
    summary_ws["B7"] = int(summary_ws["B7"].value or 0) if source_previously_listed else int(summary_ws["B7"].value or 0) + 1
    summary_ws["B9"] = excel_rows
    summary_ws["B10"] = json_rows
    summary_ws["B11"] = student_total
    summary_ws["B15"] = "学生画像与人才候选池已拆分为双视图；crawl_results 导出统一进入候选池视图。"
    summary_ws["B16"] = candidate_count
    summary_ws["B17"] = student_total + candidate_count


def update_mapping_sheet(mapping_ws, source_file_name: str) -> None:
    mapping_ws["A1"] = "字段映射说明"
    mapping_ws["A2"] = "说明：本页展示学生画像视图与人才候选池视图的字段归并规则，便于追溯映射逻辑。"
    mapping_ws["A4"] = "来源文件"

    mappings = [
        ("候选人姓名", "candidate_name/github_login"),
        ("学校", "university"),
        ("院系", "department"),
        ("邮箱", "email"),
        ("信号名称", "competition_name/venue/repo_full_name"),
        ("信号全称", "paper_title/evidence_title"),
        ("级别/排序", "award_level/ranking/author_order/contributions/followers"),
        ("时间", "season_year/venue_year/executed_at"),
        ("方向", "track"),
        ("来源URL", "evidence_url"),
        ("来源名称", "source_name"),
        ("备注", "source_id/confidence/team_name/dblp_pid/company/blog"),
    ]

    existing = {
        (
            clean_text(mapping_ws.cell(row, 1).value),
            clean_text(mapping_ws.cell(row, 2).value),
            clean_text(mapping_ws.cell(row, 3).value),
        )
        for row in range(5, mapping_ws.max_row + 1)
        if any(clean_text(mapping_ws.cell(row, col).value) for col in range(1, 4))
    }
    template_row = max(5, mapping_ws.max_row)
    for excel_field, source_field in mappings:
        record = (source_file_name, excel_field, source_field)
        if record in existing:
            continue
        target_row = mapping_ws.max_row + 1
        copy_row_style(mapping_ws, template_row, target_row, 3)
        mapping_ws.cell(target_row, 1).value = source_file_name
        mapping_ws.cell(target_row, 2).value = excel_field
        mapping_ws.cell(target_row, 3).value = source_field
        existing.add(record)
    resize_table(mapping_ws, "FieldMapping", f"A4:C{max(5, mapping_ws.max_row)}")


def resize_summary_tables(summary_ws) -> None:
    resize_table(summary_ws, "SummaryStats", "A5:B17")
    resize_table(summary_ws, "SourceCounts", "E5:H10")
    resize_table(summary_ws, "UnmappedHints", "I5:J14")


def main() -> None:
    parser = argparse.ArgumentParser(description="Split talent report into student and candidate-pool views.")
    parser.add_argument("--source", default=None, help="Path to crawl_results workbook; defaults to latest in data/exports")
    parser.add_argument("--target", default=str(DEFAULT_TARGET), help="Path to target summary workbook")
    args = parser.parse_args()

    source_path = Path(args.source).resolve() if args.source else latest_crawl_workbook(EXPORT_DIR)
    target_path = Path(args.target).resolve()

    candidate_rows, stats = collect_candidate_rows(source_path)
    workbook = load_workbook(target_path)
    student_ws = workbook[STUDENT_SHEET]
    candidate_ws = ensure_candidate_sheet(workbook, student_ws)
    mapping_ws = workbook[MAPPING_SHEET]
    summary_ws = workbook["01-执行摘要"]

    student_rows = [row for row in read_detail_rows(student_ws) if not is_crawl_result_row(row)]

    student_ws["A1"] = "人才画像汇总明细"
    student_ws["A2"] = "说明：本页保留学生画像主域数据；crawl_results 导出的候选人才信号已拆分至独立候选池视图。"
    for col, header in enumerate(STUDENT_HEADERS, start=1):
        student_ws.cell(4, col).value = header

    write_rows(student_ws, student_rows, len(STUDENT_HEADERS))
    write_rows(candidate_ws, candidate_rows, len(CANDIDATE_HEADERS))
    rebuild_table(student_ws, table_name="DetailData", headers=STUDENT_HEADERS, max_col=len(STUDENT_HEADERS), row_count=len(student_rows))
    rebuild_table(candidate_ws, table_name="CandidatePoolData", headers=CANDIDATE_HEADERS, max_col=len(CANDIDATE_HEADERS), row_count=len(candidate_rows))
    update_summary_sheet(
        summary_ws,
        source_path=source_path,
        candidate_count=len(candidate_rows),
        student_total=len(student_rows),
        source_total_rows=stats["source_total_rows"],
    )
    resize_summary_tables(summary_ws)
    update_mapping_sheet(mapping_ws, source_path.name)

    workbook.save(target_path)

    print(
        {
            "source_path": str(source_path),
            "target_path": str(target_path),
            "student_rows": len(student_rows),
            "candidate_rows": len(candidate_rows),
            "source_total_rows": stats["source_total_rows"],
            "skipped_non_structured": stats["skipped_non_structured"],
            "skipped_missing_name": stats["skipped_missing_name"],
            "source_deduped": stats["source_deduped"],
            "sheet_names": workbook.sheetnames,
        }
    )


if __name__ == "__main__":
    main()
