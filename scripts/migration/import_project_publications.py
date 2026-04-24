from __future__ import annotations

import argparse
import asyncio
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import settings
from app.db.pool import close_pool, get_pool, init_pool
from app.services.publication_service import create_formal_publication

DEFAULT_XLSX_PATH = (
    Path(__file__).resolve().parents[3]
    / "Scholars-System"
    / "docs"
    / "各项目组论文及成果产出汇总登记.xlsx"
)
DEFAULT_REPORT_PATH = (
    Path(__file__).resolve().parents[2]
    / "output"
    / "project_publication_import_report.json"
)


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def split_person_names(raw: str) -> list[str]:
    return [token.strip() for token in re.split(r"[、，,;/；]+", raw) if token.strip()]


def parse_student_tokens(raw: str) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    for part in re.split(r"[；;]+", raw):
        token = part.strip()
        if not token or "示例" in token:
            continue
        match = re.match(r"([^（(]+)[（(]?([^）)]*)", token)
        if not match:
            continue
        name = match.group(1).strip()
        student_no = match.group(2).strip()
        digits = re.search(r"\d{6,}", student_no)
        items.append(
            {
                "name": name,
                "student_no": digits.group(0) if digits else student_no,
            }
        )
    return items


def parse_authors(raw: str) -> list[str]:
    if not raw:
        return []
    if ";" in raw or "；" in raw:
        chunks = re.split(r"[；;]+", raw)
    else:
        chunks = re.split(r",\s*|\n+", raw)
    return [token.strip() for token in chunks if token.strip()]


def extract_doi(link: str) -> str | None:
    if not link:
        return None
    for pattern in (
        r"doi\.org/(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)",
        r"\b(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)\b",
    ):
        if match := re.search(pattern, link, flags=re.IGNORECASE):
            return match.group(1).rstrip(").,;")
    return None


def extract_arxiv_id(link: str) -> str | None:
    if not link:
        return None
    if match := re.search(r"arxiv\.org/(?:abs|pdf)/([0-9]{4}\.[0-9]{4,5}(?:v\d+)?)", link, flags=re.IGNORECASE):
        return match.group(1)
    return None


def infer_publication_date(*values: str) -> str | None:
    for value in values:
        if not value:
            continue
        if match := re.search(r"\b(20\d{2})\b", value):
            return f"{match.group(1)}-01-01T00:00:00+00:00"
    return None


def scholar_score(row: dict[str, Any]) -> tuple[int, str, str]:
    score = 0
    if row.get("is_academician"):
        score += 100
    for field, weight in (
        ("email", 8),
        ("profile_url", 6),
        ("google_scholar_url", 5),
        ("dblp_url", 5),
        ("university", 3),
        ("department", 2),
    ):
        if clean_text(row.get(field)):
            score += weight
    for field in ("publications_count", "h_index", "citations_count"):
        try:
            if int(row.get(field) or 0) > 0:
                score += 2
        except Exception:
            pass
    updated_at = clean_text(row.get("updated_at"))
    return score, updated_at, clean_text(row.get("id"))


async def load_match_maps() -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    pool = get_pool()
    scholar_rows = await pool.fetch(
        """
        SELECT
          id,
          name,
          university,
          department,
          email,
          profile_url,
          google_scholar_url,
          dblp_url,
          publications_count,
          h_index,
          citations_count,
          is_academician,
          updated_at
        FROM scholars
        """
    )
    student_rows = await pool.fetch(
        """
        SELECT id, name, student_no, scholar_id
        FROM supervised_students
        """
    )

    scholars_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scholar_rows:
        scholars_by_name[clean_text(row["name"])].append(dict(row))
    scholar_map: dict[str, dict[str, Any]] = {}
    for name, items in scholars_by_name.items():
        scholar_map[name] = sorted(items, key=scholar_score, reverse=True)[0]

    students_by_no: dict[str, dict[str, Any]] = {}
    students_by_name: dict[str, dict[str, Any]] = {}
    for row in student_rows:
        data = dict(row)
        if clean_text(row["student_no"]):
            students_by_no[clean_text(row["student_no"])] = data
        if clean_text(row["name"]) and clean_text(row["name"]) not in students_by_name:
            students_by_name[clean_text(row["name"])] = data

    return scholar_map, students_by_no, scholars_by_name


def iter_publication_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    ws = load_workbook(xlsx_path, data_only=True)["项目论文产出"]
    carry = {"project_batch": "", "project_name": "", "project_leader": ""}
    items: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if clean_text(row[0]):
            carry["project_batch"] = clean_text(row[0])
        if clean_text(row[1]):
            carry["project_name"] = clean_text(row[1])
        if clean_text(row[2]):
            carry["project_leader"] = clean_text(row[2])

        title = clean_text(row[4])
        if not title:
            continue

        venue = clean_text(row[5])
        link = clean_text(row[10])
        items.append(
            {
                "row_index": row_index,
                "project_batch": carry["project_batch"],
                "project_name": carry["project_name"],
                "project_leader": carry["project_leader"],
                "title": title,
                "venue": venue,
                "authors": parse_authors(clean_text(row[6])),
                "student_tokens": parse_student_tokens(clean_text(row[7])),
                "unit_acknowledged": clean_text(row[8]),
                "paper_status": clean_text(row[9]),
                "article_link": link,
                "abstract": clean_text(row[11]),
                "note": clean_text(row[12]),
                "doi": extract_doi(link),
                "arxiv_id": extract_arxiv_id(link),
                "publication_date": infer_publication_date(venue, link, clean_text(row[9]), title),
            }
        )
    return items


async def main() -> None:
    parser = argparse.ArgumentParser(description="Import project publications from Excel into unified publication tables.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH))
    parser.add_argument("--report", default=str(DEFAULT_REPORT_PATH))
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    report_path = Path(args.report).resolve()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    await init_pool(
        host=settings.POSTGRES_HOST,
        port=settings.POSTGRES_PORT,
        user=settings.POSTGRES_USER,
        password=settings.POSTGRES_PASSWORD,
        database=settings.POSTGRES_DB,
    )

    try:
        scholar_map, students_by_no, scholars_by_name = await load_match_maps()
        rows = iter_publication_rows(xlsx_path)

        report: dict[str, Any] = {
            "xlsx_path": str(xlsx_path),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {
                "rows_seen": len(rows),
                "publications_processed": 0,
                "owner_links_upserted": 0,
                "scholar_links": 0,
                "student_links": 0,
                "rows_without_owner": 0,
            },
            "unmatched_leaders": [],
            "unmatched_students": [],
            "resolved_ambiguous_leaders": [],
        }

        unmatched_leader_keys: set[tuple[int, str]] = set()
        unmatched_student_keys: set[tuple[int, str, str]] = set()

        for item in rows:
            owners: list[tuple[str, str, dict[str, Any]]] = []

            for raw_name in split_person_names(item["project_leader"]):
                resolved = scholar_map.get(raw_name)
                if resolved:
                    owners.append(("scholar", clean_text(resolved["id"]), {"match_name": raw_name}))
                    if len(scholars_by_name.get(raw_name, [])) > 1:
                        report["resolved_ambiguous_leaders"].append(
                            {
                                "row_index": item["row_index"],
                                "leader_name": raw_name,
                                "selected_scholar_id": clean_text(resolved["id"]),
                                "candidate_ids": [clean_text(row["id"]) for row in scholars_by_name.get(raw_name, [])],
                            }
                        )
                elif (item["row_index"], raw_name) not in unmatched_leader_keys:
                    unmatched_leader_keys.add((item["row_index"], raw_name))
                    report["unmatched_leaders"].append(
                        {
                            "row_index": item["row_index"],
                            "leader_name": raw_name,
                            "project_name": item["project_name"],
                            "title": item["title"],
                        }
                    )

            for token in item["student_tokens"]:
                student_row = None
                if clean_text(token["student_no"]):
                    student_row = students_by_no.get(clean_text(token["student_no"]))
                if student_row is None and clean_text(token["name"]):
                    # fallback by exact name only when student_no missing
                    pool = get_pool()
                    student_row = await pool.fetchrow(
                        """
                        SELECT id, name, student_no, scholar_id
                        FROM supervised_students
                        WHERE name = $1
                        ORDER BY updated_at DESC NULLS LAST, created_at DESC, id ASC
                        LIMIT 1
                        """,
                        clean_text(token["name"]),
                    )
                    student_row = dict(student_row) if student_row else None
                if student_row:
                    owners.append(("student", clean_text(student_row["id"]), token))
                elif (item["row_index"], token["name"], token["student_no"]) not in unmatched_student_keys:
                    unmatched_student_keys.add((item["row_index"], token["name"], token["student_no"]))
                    report["unmatched_students"].append(
                        {
                            "row_index": item["row_index"],
                            "student_name": token["name"],
                            "student_no": token["student_no"],
                            "project_name": item["project_name"],
                            "title": item["title"],
                        }
                    )

            deduped_owners: dict[tuple[str, str], dict[str, Any]] = {}
            for owner_type, owner_id, meta in owners:
                deduped_owners[(owner_type, owner_id)] = meta
            owners = [(owner_type, owner_id, meta) for (owner_type, owner_id), meta in deduped_owners.items()]

            if not owners:
                report["summary"]["rows_without_owner"] += 1
                continue

            report["summary"]["publications_processed"] += 1
            for owner_type, owner_id, meta in owners:
                source_details = {
                    "file_name": xlsx_path.name,
                    "sheet_name": "项目论文产出",
                    "row_index": item["row_index"],
                    "project_batch": item["project_batch"],
                    "project_name": item["project_name"],
                    "project_leader": item["project_leader"],
                    "venue": item["venue"],
                    "article_link": item["article_link"],
                    "paper_status": item["paper_status"],
                    "unit_acknowledged": item["unit_acknowledged"],
                    "note": item["note"],
                    "owner_match": meta,
                    "import_channel": "project_excel",
                }
                compliance_details = {
                    "paper_status": item["paper_status"],
                    "unit_acknowledged": item["unit_acknowledged"],
                    "note": item["note"],
                    "imported_from_excel": True,
                }
                await create_formal_publication(
                    get_pool(),
                    owner_type=owner_type,
                    owner_id=owner_id,
                    title=item["title"],
                    doi=item["doi"],
                    arxiv_id=item["arxiv_id"],
                    abstract=item["abstract"] or None,
                    publication_date=item["publication_date"],
                    authors=item["authors"],
                    affiliations=[],
                    project_group_name=item["project_name"] or None,
                    source_type="bulk_import",
                    source_details=source_details,
                    compliance_details=compliance_details,
                    confirmed_by="system:excel_import",
                )
                report["summary"]["owner_links_upserted"] += 1
                if owner_type == "scholar":
                    report["summary"]["scholar_links"] += 1
                else:
                    report["summary"]["student_links"] += 1

        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(report["summary"], ensure_ascii=False))
        print(f"report_saved={report_path}")
    finally:
        await close_pool()


if __name__ == "__main__":
    asyncio.run(main())
