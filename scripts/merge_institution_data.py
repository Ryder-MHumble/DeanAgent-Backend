#!/usr/bin/env python3
"""
Merge Excel collaboration data into existing institution.json file.

Usage:
    python scripts/merge_institution_data.py
"""
import json
from pathlib import Path

from openpyxl import load_workbook


def parse_multiline_list(text: str | None) -> list[str]:
    """Parse multiline text into list of items."""
    if not text:
        return []
    lines = [line.strip() for line in str(text).split("\n") if line.strip()]
    return lines


def parse_person_list(text: str | None, role_key: str = "role") -> list[dict]:
    """Parse person information from multiline text."""
    if not text:
        return []

    lines = parse_multiline_list(text)
    persons = []

    for line in lines:
        # Extract name (first line or before first newline/colon)
        parts = line.split("\n")
        name = parts[0].strip()

        person = {"name": name}
        if len(parts) > 1:
            person[role_key] = "\n".join(parts[1:]).strip()

        persons.append(person)

    return persons


def clean_priority(priority: str | None) -> tuple[str, bool]:
    """Clean priority field and extract demo school flag."""
    if not priority:
        return "", False

    priority_str = str(priority).strip()
    is_demo = "示范校" in priority_str

    # Extract P0/P1/P2/P3
    import re
    match = re.search(r"P[0-3]", priority_str)
    if match:
        return match.group(0), is_demo

    return "", is_demo


def main():
    # Load existing institution.json
    institution_file = Path("data/institution.json")
    with open(institution_file, "r", encoding="utf-8") as f:
        existing_data = json.load(f)

    # Create name to org mapping
    org_map = {}
    for org in existing_data["organizations"]:
        org_map[org["name_zh"]] = org

    print(f"Loaded {len(org_map)} existing organizations")

    # Load Excel data
    excel_path = Path("docs/共建高校信息汇总.xlsx")
    print(f"\nLoading Excel file: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Build headers from row 5 and 6
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        main_header = ws.cell(row=5, column=col_idx).value
        sub_header = ws.cell(row=6, column=col_idx).value

        if main_header and sub_header:
            header = f"{main_header}_{sub_header}"
        elif main_header:
            header = main_header
        elif sub_header:
            header = sub_header
        else:
            header = f"col_{col_idx}"

        headers.append(header)

    # Extract and merge data
    merged_count = 0
    new_count = 0

    for row_idx in range(7, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[header] = cell.value

        # Skip empty rows
        university_name = row_data.get("高校")
        if not university_name:
            continue

        # Parse priority and demo school flag
        priority, is_demo = clean_priority(row_data.get("优先级"))

        # Build collaboration data
        collaboration_data = {
            # Classification and priority
            "category": row_data.get("分类") or "",
            "priority": priority,
            "is_demo_school": is_demo,

            # Student and supervisor counts
            "student_count_24": int(row_data.get("学生人数_24级") or 0),
            "student_count_25": int(row_data.get("25级") or 0),
            "student_count_total": int(row_data.get("总数") or 0),
            "supervisor_count": int(row_data.get("导师人数_总数") or 0),

            # Personnel (structured)
            "institute_leaders": parse_person_list(
                row_data.get("驻院领导及共建老师"), "roles"
            ),
            "degree_committee": parse_person_list(
                row_data.get("学位委员"), "role"
            ),
            "teaching_committee": parse_person_list(
                row_data.get("教学委员"), "role"
            ),
            "joint_supervisors": parse_person_list(
                row_data.get("共建导师_姓名"), "info"
            ),
            "university_leaders": parse_person_list(
                row_data.get("相关校领导_姓名"), "title"
            ),
            "important_scholars": parse_person_list(
                row_data.get("重要学者_姓名"), "title"
            ),

            # Departments and collaboration
            "key_departments": parse_multiline_list(row_data.get("重点院系")),
            "collaboration": {
                "joint_lab": row_data.get("合作_联合实验室") or "",
                "training": row_data.get("培养") or "",
                "academic": row_data.get("学术") or "",
                "talent_hiring": row_data.get("人才双聘") or "",
            },
            "collaboration_focus": row_data.get("合作重点") or "",

            # Exchange records
            "exchange_records": [
                {"event": row_data.get("交流互访_招生宣讲") or ""},
                {"event": row_data.get("col_27") or ""},
            ],

            # Metadata
            "data_source": "共建高校信息汇总.xlsx (2025-07-18)",
            "last_updated": "2025-07-18",
        }

        # Remove empty exchange records
        collaboration_data["exchange_records"] = [
            r for r in collaboration_data["exchange_records"] if r["event"]
        ]

        # Merge with existing organization
        if university_name in org_map:
            org_map[university_name].update(collaboration_data)
            merged_count += 1
            print(f"  ✓ Merged: {university_name}")
        else:
            # Add new organization
            new_org = {
                "name_zh": university_name,
                "name_en": "",
                "org_id": "",
                "org_name": "",
            }
            new_org.update(collaboration_data)
            existing_data["organizations"].append(new_org)
            org_map[university_name] = new_org
            new_count += 1
            print(f"  + Added: {university_name}")

    # Save merged data
    with open(institution_file, "w", encoding="utf-8") as f:
        json.dump(existing_data, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"✓ Successfully merged data to {institution_file}")
    print(f"{'='*60}")
    print(f"  Merged existing: {merged_count}")
    print(f"  Added new: {new_count}")
    print(f"  Total organizations: {len(existing_data['organizations'])}")

    # Print summary statistics
    categories = {}
    priorities = {}
    demo_count = 0
    total_students = 0
    total_supervisors = 0

    for org in existing_data["organizations"]:
        if "category" in org:
            cat = org.get("category") or "Unknown"
            categories[cat] = categories.get(cat, 0) + 1

            pri = org.get("priority") or "Unknown"
            priorities[pri] = priorities.get(pri, 0) + 1

            if org.get("is_demo_school"):
                demo_count += 1

            total_students += org.get("student_count_total", 0)
            total_supervisors += org.get("supervisor_count", 0)

    print(f"\n{'='*60}")
    print("STATISTICS")
    print(f"{'='*60}")
    print(f"  Demo schools: {demo_count}")
    print(f"  Total students: {total_students}")
    print(f"  Total supervisors: {total_supervisors}")

    if categories:
        print(f"\n  By category:")
        for cat, count in sorted(categories.items()):
            print(f"    {cat}: {count}")

    if priorities:
        print(f"\n  By priority:")
        for pri, count in sorted(priorities.items()):
            print(f"    {pri}: {count}")


if __name__ == "__main__":
    main()
