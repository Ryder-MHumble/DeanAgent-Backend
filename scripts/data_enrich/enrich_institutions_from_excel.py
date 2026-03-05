#!/usr/bin/env python3
"""
从共建高校信息汇总.xlsx 补充 institutions.json 的详情数据
"""
import json
import sys
from pathlib import Path
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def parse_multiline_field(text: str | None) -> list[str]:
    """解析多行文本字段（用换行符分隔）"""
    if not text:
        return []
    return [line.strip() for line in str(text).split('\n') if line.strip()]


def parse_scholars(names: str | None, titles: str | None, departments: str | None,
                   research_areas: str | None) -> list[dict]:
    """解析学者信息（姓名、头衔、院系、研究方向）"""
    name_list = parse_multiline_field(names)
    title_list = parse_multiline_field(titles)
    dept_list = parse_multiline_field(departments)
    research_list = parse_multiline_field(research_areas)

    # 对齐长度
    max_len = max(len(name_list), len(title_list), len(dept_list), len(research_list))
    name_list += [''] * (max_len - len(name_list))
    title_list += [''] * (max_len - len(title_list))
    dept_list += [''] * (max_len - len(dept_list))
    research_list += [''] * (max_len - len(research_list))

    scholars = []
    for name, title, dept, research in zip(name_list, title_list, dept_list, research_list):
        if name:
            scholars.append({
                "name": name,
                "title": title or None,
                "department": dept or None,
                "research_area": research or None
            })

    return scholars


def parse_excel(excel_path: Path) -> dict[str, dict]:
    """解析 Excel 文件，返回 {高校名: 详情数据}"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 第5行是主表头，第6行是子表头
    main_headers = [cell.value for cell in ws[5]]
    sub_headers = [cell.value for cell in ws[6]]

    # 合并表头
    headers = []
    for main, sub in zip(main_headers, sub_headers):
        if sub:
            headers.append(sub)
        elif main:
            headers.append(main)
        else:
            headers.append(None)

    # 解析数据行（从第7行开始）
    universities = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row[2]:  # 高校名为空，跳过
            continue

        univ_name = str(row[2]).strip()

        # 提取字段（按列索引）
        data = {
            "category": row[1],  # 分类
            "student_count_24": row[3] if isinstance(row[3], int) else None,  # 24级学生数
            "student_count_25": row[4] if isinstance(row[4], int) else None,  # 25级学生数
            "student_count_total": row[5] if isinstance(row[5], int) else None,  # 学生总数
            "mentor_count": row[6] if isinstance(row[6], int) else None,  # 导师总数
            "resident_leaders": parse_multiline_field(row[7]),  # 驻院领导及共建老师
            "degree_committee": parse_multiline_field(row[8]),  # 学位委员
            "teaching_committee": parse_multiline_field(row[9]),  # 教学委员
            "mentor_info": {
                "name": row[10],
                "category": row[11],
                "department": row[12]
            } if row[10] else None,
            "university_leaders": parse_scholars(row[13], row[14], None, row[15]),  # 相关校领导
            "notable_scholars": parse_scholars(row[16], row[17], row[18], row[19]),  # 重要学者
            "key_departments": parse_multiline_field(row[20]),  # 重点院系
            "joint_labs": parse_multiline_field(row[21]),  # 联合实验室
            "training_cooperation": parse_multiline_field(row[22]),  # 培养合作
            "academic_cooperation": parse_multiline_field(row[23]),  # 学术合作
            "talent_dual_appointment": parse_multiline_field(row[24]),  # 人才双聘
            "recruitment_events": parse_multiline_field(row[25]),  # 招生宣讲
            "visit_exchanges": parse_multiline_field(row[26]) if len(row) > 26 else [],  # 交流互访
            "priority": row[27] if len(row) > 27 else None,  # 优先级
            "cooperation_focus": parse_multiline_field(row[28]) if len(row) > 28 else []  # 合作重点
        }

        universities[univ_name] = data

    return universities


def enrich_institutions(institutions_path: Path, excel_path: Path) -> None:
    """补充 institutions.json 数据"""
    # 读取现有 institutions.json
    with open(institutions_path, 'r', encoding='utf-8') as f:
        institutions_data = json.load(f)

    # 解析 Excel
    excel_data = parse_excel(excel_path)

    print(f"从 Excel 解析到 {len(excel_data)} 所高校的详情数据")

    # 补充数据
    enriched_count = 0
    for univ in institutions_data['universities']:
        univ_name = univ['name']

        if univ_name in excel_data:
            # 合并 Excel 数据
            univ['details'] = excel_data[univ_name]
            enriched_count += 1
            print(f"✓ 补充 {univ_name} 的详情数据")
        else:
            print(f"⚠ {univ_name} 在 Excel 中未找到")

    # 更新时间戳
    institutions_data['last_updated'] = datetime.now(timezone.utc).isoformat()

    # 写回文件
    with open(institutions_path, 'w', encoding='utf-8') as f:
        json.dump(institutions_data, f, ensure_ascii=False, indent=2)

    print(f"\n✓ 成功补充 {enriched_count}/{len(institutions_data['universities'])} 所高校的详情数据")
    print(f"✓ 已更新 {institutions_path}")


def main():
    project_root = Path(__file__).parent.parent
    institutions_path = project_root / 'data' / 'institutions.json'
    excel_path = project_root / 'docs' / '共建高校信息汇总.xlsx'

    if not institutions_path.exists():
        print(f"Error: {institutions_path} 不存在")
        sys.exit(1)

    if not excel_path.exists():
        print(f"Error: {excel_path} 不存在")
        sys.exit(1)

    enrich_institutions(institutions_path, excel_path)


if __name__ == '__main__':
    main()
