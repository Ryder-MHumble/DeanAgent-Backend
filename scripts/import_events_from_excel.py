#!/usr/bin/env python3
"""从 Excel 导入活动数据到 events.json"""
import json
import sys
import uuid
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q", "--break-system-packages"])
    import openpyxl


def import_events_from_excel(excel_path: str, output_path: str = "data/events.json"):
    """从讲座信息汇总 Excel 导入活动数据"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 读取表头（第1行）
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    events = []

    # 从第2行开始读取数据
    for row_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]

        # 跳过空行
        if not any(row):
            continue

        # 映射字段（根据实际 Excel 结构）
        try:
            event_data = {
                "id": str(uuid.uuid4()),
                "event_type": row[8] if len(row) > 8 and row[8] else "讲座",  # 讲座类型
                "series_number": str(row[9]) if len(row) > 9 and row[9] else "",  # 系列期数

                # 讲者信息
                "speaker_name": row[0] if len(row) > 0 and row[0] else "",  # 姓名
                "speaker_organization": row[1] if len(row) > 1 and row[1] else "",  # 工作单位
                "speaker_position": row[2] if len(row) > 2 and row[2] else "",  # 职务
                "speaker_bio": row[3] if len(row) > 3 and row[3] else "",  # 讲者简介
                "speaker_photo_url": row[4] if len(row) > 4 and row[4] else "",  # 讲者照片

                # 活动信息
                "title": row[10] if len(row) > 10 and row[10] else "",  # 讲座题目
                "abstract": row[11] if len(row) > 11 and row[11] else "",  # 讲座摘要
                "event_date": _parse_date(row[5]) if len(row) > 5 else "",  # 讲座日期
                "duration": float(row[6]) if len(row) > 6 and row[6] and isinstance(row[6], (int, float)) else 1.0,  # 讲座时长
                "location": row[7] if len(row) > 7 and row[7] else "",  # 地点

                # 关联信息
                "scholar_ids": [],

                # 管理信息
                "publicity": row[12] if len(row) > 12 and row[12] else "",  # 公关传播
                "needs_email_invitation": bool(row[13]) if len(row) > 13 and row[13] else False,  # 需要邮件邀请
                "certificate_number": str(row[14]) if len(row) > 14 and row[14] and row[14] != "#N/A" else "",  # 纪念证书编号
                "created_by": row[15] if len(row) > 15 and row[15] else "system",  # 创建人
                "created_at": _parse_date(row[16]) if len(row) > 16 else datetime.now().isoformat(),  # 创建时间
                "updated_at": datetime.now().isoformat(),
                "audit_status": row[17] if len(row) > 17 and row[17] else "pending",  # 审核流程
            }

            # 只添加有效的活动（至少有讲者姓名或题目）
            if event_data["speaker_name"] or event_data["title"]:
                events.append(event_data)
                print(f"  ✓ Row {row_idx}: {event_data['speaker_name']} - {event_data['title'][:30]}...")

        except Exception as e:
            print(f"  ✗ Row {row_idx}: Error - {e}")
            continue

    # 保存到 JSON
    output_data = {
        "total": len(events),
        "last_updated": datetime.now().isoformat(),
        "events": events
    }

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"✓ Imported {len(events)} events to {output_path}")
    print(f"{'='*60}")


def _parse_date(value) -> str:
    """解析日期字段"""
    if not value:
        return ""

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, str):
        # 尝试解析常见日期格式
        for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y/%m/%d", "%Y.%m.%d"]:
            try:
                dt = datetime.strptime(value, fmt)
                return dt.isoformat()
            except ValueError:
                continue
        return value

    return str(value)


if __name__ == "__main__":
    excel_file = "docs/讲座信息汇总.xlsx"

    if not Path(excel_file).exists():
        print(f"✗ Excel file not found: {excel_file}")
        sys.exit(1)

    import_events_from_excel(excel_file)
