from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook

from app.crawlers.base import CrawledItem, CrawlResult, CrawlStatus
from app.schemas.console import CrawlRequest
from app.services.crawler_control_service import CrawlerControlService


class _FakeCrawler:
    def __init__(self, source_config: dict[str, Any], domain_keywords: list[str] | None = None):
        self.source_config = source_config
        self.domain_keywords = domain_keywords

    async def run(self) -> CrawlResult:
        now = datetime(2026, 4, 24, 12, 0, tzinfo=timezone.utc)
        item = CrawledItem(
            title="Alice Chen",
            url="https://example.com/kaggle/alice",
            source_id=self.source_config["id"],
            dimension="talent_scout",
            extra={
                "competition_name": "Kaggle",
                "season_year": 2026,
                "award_level": "Gold",
                "ranking": "1",
                "team_name": "Solo",
                "talent_signal": {
                    "signal_type": "competition",
                    "candidate_name": "Alice Chen",
                    "university": "浙江大学",
                    "department": "计算机学院",
                    "email": "",
                    "track": "ml",
                    "record_status": "structured",
                    "confidence": 0.99,
                    "identity_hints": {"github": "alice"},
                    "source_metrics": {"ranking": 1},
                    "evidence_title": "Kaggle Ranking",
                    "evidence_url": "https://example.com/kaggle/alice",
                    "notes": "",
                },
            },
        )
        return CrawlResult(
            source_id=self.source_config["id"],
            status=CrawlStatus.SUCCESS,
            items=[item],
            items_all=[item],
            items_new=1,
            items_total=1,
            started_at=now,
            finished_at=now,
            duration_seconds=0.0,
        )


@pytest.mark.asyncio
async def test_crawl_request_accepts_xlsx_export_format():
    payload = CrawlRequest(source_ids=["kaggle_grandmaster"], export_format="xlsx")

    assert payload.export_format == "xlsx"


@pytest.mark.asyncio
async def test_crawler_control_service_exports_xlsx_without_database_persistence(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    monkeypatch.setattr(
        "app.services.crawler_control_service.BASE_DIR",
        tmp_path,
    )
    monkeypatch.setattr(
        "app.services.crawler_control_service.load_all_source_configs",
        lambda: [
            {
                "id": "kaggle_grandmaster",
                "name": "Kaggle Grandmaster",
                "dimension": "talent_scout",
                "entity_family": "competition",
                "capture_mode": "structured",
                "fallback_mode": "evidence_only",
                "adapter_key": "kaggle",
                "sheet_name": "KaggleGM",
                "seed_urls": ["https://example.com/kaggle"],
                "crawl_method": "static",
                "schedule": "daily",
                "is_enabled": False,
                "requires_auth": False,
                "crawler_class": "competition_source",
                "tracks": ["ml"],
            }
        ],
    )
    monkeypatch.setattr(
        "app.services.crawler_control_service.create_crawler",
        lambda config, domain_keywords=None: _FakeCrawler(config, domain_keywords),
    )
    append_log = AsyncMock()
    update_state = AsyncMock()
    save_db = AsyncMock()
    monkeypatch.setattr("app.services.crawler_control_service.append_crawl_log", append_log)
    monkeypatch.setattr("app.services.crawler_control_service.update_source_state", update_state)
    monkeypatch.setattr("app.services.crawler_control_service.save_crawl_result_json", save_db)

    service = CrawlerControlService()
    job = service.create_job(
        source_ids=["kaggle_grandmaster"],
        export_format="xlsx",
    )

    await service._jobs[job["job_id"]]["task"]
    final_job = service.get_job(job["job_id"])

    assert final_job["status"] == "completed"
    assert final_job["result_file_name"] is not None
    assert final_job["result_file_name"].endswith(".xlsx")
    assert save_db.await_count == 0
    assert append_log.await_count == 1
    assert update_state.await_count == 1

    file_path = service.get_job_result_file(job["job_id"])
    assert file_path is not None
    workbook = load_workbook(file_path)
    assert workbook.sheetnames == ["README", "SourceStatus", "TalentSignals"]
