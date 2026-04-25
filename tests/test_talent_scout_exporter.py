from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import load_workbook

from app.services.talent_scout.excel_exporter import export_talent_scout_workbook


def test_export_talent_scout_workbook_writes_status_and_single_signal_sheet(tmp_path):
    executed_at = datetime(2026, 4, 24, 12, 0, tzinfo=timezone.utc)
    source_configs = [
        {
            "id": "kaggle_grandmaster",
            "name": "Kaggle Grandmaster",
            "entity_family": "competition",
            "capture_mode": "structured",
            "sheet_name": "KaggleGM",
            "requires_auth": False,
            "adapter_key": "kaggle",
            "seed_urls": ["https://example.com/kaggle"],
        },
        {
            "id": "aliyun_tianchi",
            "name": "阿里天池",
            "entity_family": "competition",
            "capture_mode": "structured",
            "sheet_name": "Tianchi",
            "requires_auth": True,
            "adapter_key": "tianchi",
            "seed_urls": ["https://example.com/tianchi"],
        },
    ]
    source_runs = [
        {
            "source_id": "kaggle_grandmaster",
            "status": "success",
            "items_dict": [
                {
                    "source_id": "kaggle_grandmaster",
                    "title": "Alice Chen",
                    "url": "https://example.com/kaggle/alice",
                    "extra": {
                        "competition_name": "Kaggle",
                        "season_year": 2026,
                        "award_level": "Gold",
                        "ranking": "1",
                        "team_name": "Solo",
                        "talent_signal": {
                            "signal_type": "competition",
                            "candidate_name": "Alice Chen",
                            "university": "浙江大学",
                            "department": "计算机学院",
                            "email": "",
                            "track": "ml",
                            "record_status": "structured",
                            "confidence": 0.98,
                            "identity_hints": {"github": "alice"},
                            "source_metrics": {"ranking": 1},
                            "evidence_title": "Kaggle Ranking",
                            "evidence_url": "https://example.com/kaggle/alice",
                            "notes": "",
                        },
                    },
                }
            ],
            "executed_at": executed_at,
            "error_message": None,
        },
        {
            "source_id": "aliyun_tianchi",
            "status": "partial",
            "items_dict": [],
            "executed_at": executed_at,
            "error_message": "auth_required",
        },
    ]

    file_path = export_talent_scout_workbook(
        source_configs=source_configs,
        source_runs=source_runs,
        output_dir=tmp_path,
        generated_at=executed_at,
    )

    workbook = load_workbook(file_path)

    assert workbook.sheetnames == ["README", "SourceStatus", "TalentSignals"]

    source_status = workbook["SourceStatus"]
    assert [cell.value for cell in source_status[1]] == [
        "source_id",
        "source_name",
        "entity_family",
        "capture_mode",
        "crawl_status",
        "rows_exported",
        "requires_auth",
        "block_reason",
        "adapter_key",
        "executed_at",
    ]
    assert source_status.cell(2, 1).value == "kaggle_grandmaster"
    assert source_status.cell(3, 1).value == "aliyun_tianchi"
    assert source_status.cell(3, 5).value == "blocked"
    assert source_status.cell(3, 6).value == 0
    assert source_status.cell(3, 7).value is True

    signals_sheet = workbook["TalentSignals"]
    assert [cell.value for cell in signals_sheet[1]] == [
        "candidate_name",
        "university",
        "department",
        "email",
        "track",
        "source_id",
        "source_name",
        "record_status",
        "confidence",
        "evidence_title",
        "evidence_url",
        "executed_at",
        "competition_name",
        "season_year",
        "award_level",
        "ranking",
        "team_name",
        "venue",
        "venue_year",
        "paper_title",
        "author_order",
        "paper_count_in_scope",
        "citation_count",
        "dblp_pid",
        "github_login",
        "repo_full_name",
        "contributions",
        "followers",
        "company",
        "blog",
    ]
    assert signals_sheet.cell(2, 1).value == "Alice Chen"
    assert signals_sheet.cell(2, 7).value == "Kaggle Grandmaster"
    assert signals_sheet.cell(2, 13).value == "Kaggle"
    assert signals_sheet.cell(3, 6).value == "aliyun_tianchi"
    assert signals_sheet.cell(3, 7).value == "阿里天池"
    assert signals_sheet.cell(3, 8).value == "blocked"
    assert signals_sheet.cell(3, 11).value == "https://example.com/tianchi"


def test_export_talent_scout_workbook_does_not_count_placeholder_as_exported_row(
    tmp_path,
):
    file_path = export_talent_scout_workbook(
        source_configs=[
            {
                "id": "lanqiao",
                "name": "蓝桥杯",
                "entity_family": "competition",
                "capture_mode": "semi_structured",
                "adapter_key": "rank_table",
                "requires_auth": False,
                "tracks": ["algorithm"],
                "seed_urls": ["https://dasai.lanqiao.cn/"],
            }
        ],
        source_runs=[
            {
                "source_id": "lanqiao",
                "status": "no_new_content",
                "items_dict": [],
                "executed_at": datetime(2026, 4, 24, 12, 0, tzinfo=timezone.utc),
            }
        ],
        output_dir=tmp_path,
        generated_at=datetime(2026, 4, 24, 12, 0, tzinfo=timezone.utc),
    )

    workbook = load_workbook(file_path)
    source_status = workbook["SourceStatus"]
    talent_signals = workbook["TalentSignals"]

    assert source_status.cell(2, 5).value == "no_new_content"
    assert source_status.cell(2, 6).value == 0
    assert talent_signals.cell(2, 1).value is None
    assert talent_signals.cell(2, 8).value == "needs_review"


def test_export_talent_scout_workbook_does_not_count_blocked_signal_as_candidate_row(
    tmp_path,
):
    file_path = export_talent_scout_workbook(
        source_configs=[
            {
                "id": "aliyun_tianchi",
                "name": "阿里天池",
                "entity_family": "competition",
                "capture_mode": "evidence_only",
                "adapter_key": "manual_seed",
                "requires_auth": True,
                "tracks": ["data_mining"],
                "seed_urls": ["https://tianchi.aliyun.com/competition"],
            }
        ],
        source_runs=[
            {
                "source_id": "aliyun_tianchi",
                "status": "success",
                "items_dict": [
                    {
                        "title": "阿里天池",
                        "url": "https://tianchi.aliyun.com/competition",
                        "extra": {
                            "talent_signal": {
                                "signal_type": "competition",
                                "candidate_name": "",
                                "university": "",
                                "department": "",
                                "email": "",
                                "track": "data_mining",
                                "record_status": "blocked",
                                "confidence": 0.0,
                                "identity_hints": {},
                                "source_metrics": {},
                                "evidence_title": "阿里天池",
                                "evidence_url": "https://tianchi.aliyun.com/competition",
                                "notes": "auth required",
                            }
                        },
                    }
                ],
                "executed_at": datetime(2026, 4, 24, 12, 0, tzinfo=timezone.utc),
            }
        ],
        output_dir=tmp_path,
        generated_at=datetime(2026, 4, 24, 12, 0, tzinfo=timezone.utc),
    )

    workbook = load_workbook(file_path)
    source_status = workbook["SourceStatus"]
    talent_signals = workbook["TalentSignals"]

    assert source_status.cell(2, 5).value == "blocked"
    assert source_status.cell(2, 6).value == 0
    assert source_status.cell(2, 8).value == "auth required"
    assert talent_signals.cell(2, 1).value is None
    assert talent_signals.cell(2, 8).value == "blocked"
