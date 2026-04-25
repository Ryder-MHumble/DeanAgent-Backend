from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

COMMON_SOURCE_COLUMNS = [
    "candidate_name",
    "university",
    "department",
    "email",
    "track",
    "source_id",
    "source_name",
    "record_status",
    "confidence",
    "evidence_title",
    "evidence_url",
    "executed_at",
]

COMPETITION_COLUMNS = [
    "competition_name",
    "season_year",
    "award_level",
    "ranking",
    "team_name",
]

PAPER_COLUMNS = [
    "venue",
    "venue_year",
    "paper_title",
    "author_order",
    "paper_count_in_scope",
    "citation_count",
    "dblp_pid",
]

GITHUB_COLUMNS = [
    "github_login",
    "repo_full_name",
    "contributions",
    "followers",
    "company",
    "blog",
]

TALENT_SIGNAL_COLUMNS = [
    *COMMON_SOURCE_COLUMNS,
    *COMPETITION_COLUMNS,
    *PAPER_COLUMNS,
    *GITHUB_COLUMNS,
]

SOURCE_STATUS_COLUMNS = [
    "source_id",
    "source_name",
    "entity_family",
    "capture_mode",
    "crawl_status",
    "rows_exported",
    "requires_auth",
    "block_reason",
    "adapter_key",
    "executed_at",
]


def export_talent_scout_workbook(
    *,
    source_configs: list[dict[str, Any]],
    source_runs: list[dict[str, Any]],
    output_dir: Path,
    generated_at: datetime | None = None,
) -> Path:
    timestamp = _normalize_datetime(generated_at or datetime.now(timezone.utc))
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"crawl_results_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"

    workbook = Workbook()
    workbook.remove(workbook.active)

    workbook.create_sheet("README")
    _write_readme_sheet(
        workbook["README"],
        generated_at=generated_at or timestamp,
        source_count=len(source_configs),
    )

    workbook.create_sheet("SourceStatus")
    _write_source_status_sheet(
        workbook["SourceStatus"],
        source_configs=source_configs,
        source_runs=source_runs,
    )

    workbook.create_sheet("TalentSignals")
    _write_talent_signals_sheet(
        workbook["TalentSignals"],
        source_configs=source_configs,
        source_runs=source_runs,
    )

    workbook.save(file_path)
    return file_path


def _write_readme_sheet(
    worksheet: Any,
    *,
    generated_at: datetime,
    source_count: int,
) -> None:
    rows = [
        ["Talent Scout Export", ""],
        ["generated_at", _format_datetime(generated_at)],
        ["source_count", source_count],
        ["sheets", "README, SourceStatus, TalentSignals"],
    ]
    for row in rows:
        worksheet.append(row)


def _write_source_status_sheet(
    worksheet: Any,
    *,
    source_configs: list[dict[str, Any]],
    source_runs: list[dict[str, Any]],
) -> None:
    worksheet.append(SOURCE_STATUS_COLUMNS)
    runs_by_id = {
        str(run.get("source_id") or ""): run for run in source_runs if run.get("source_id")
    }
    for config in source_configs:
        source_id = str(config.get("id") or "")
        run = runs_by_id.get(source_id, {"source_id": source_id})
        rows = _build_source_rows(config=config, run=run)
        crawl_status = _resolve_source_status(config=config, run=run, rows=rows)
        block_reason = _resolve_block_reason(run=run, rows=rows)
        worksheet.append(
            [
                source_id,
                config.get("name"),
                config.get("entity_family"),
                config.get("capture_mode"),
                crawl_status,
                _count_exported_rows(rows),
                bool(config.get("requires_auth", False)),
                block_reason,
                config.get("adapter_key"),
                _format_datetime(run.get("executed_at")),
            ]
        )


def _write_talent_signals_sheet(
    worksheet: Any,
    *,
    source_configs: list[dict[str, Any]],
    source_runs: list[dict[str, Any]],
) -> None:
    worksheet.append(TALENT_SIGNAL_COLUMNS)

    for row in _build_all_signal_rows(source_configs=source_configs, source_runs=source_runs):
        worksheet.append([row.get(column) for column in TALENT_SIGNAL_COLUMNS])


def _build_all_signal_rows(
    *,
    source_configs: list[dict[str, Any]],
    source_runs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    runs_by_id = {
        str(run.get("source_id") or ""): run for run in source_runs if run.get("source_id")
    }
    rows: list[dict[str, Any]] = []
    for config in source_configs:
        source_id = str(config.get("id") or "")
        rows.extend(
            _build_source_rows(
                config=config,
                run=runs_by_id.get(source_id, {"source_id": source_id}),
            )
        )
    return rows


def _build_source_rows(*, config: dict[str, Any], run: dict[str, Any]) -> list[dict[str, Any]]:
    source_id = str(config.get("id") or run.get("source_id") or "")
    source_name = str(config.get("name") or source_id)
    executed_at = _format_datetime(run.get("executed_at"))

    rows: list[dict[str, Any]] = []
    for item in list(run.get("items_dict") or []):
        signal = dict(item.get("extra", {}).get("talent_signal") or {})
        candidate_name = signal.get("candidate_name") if signal else item.get("title")
        row = {
            "candidate_name": candidate_name or None,
            "university": signal.get("university"),
            "department": signal.get("department"),
            "email": signal.get("email"),
            "track": signal.get("track"),
            "source_id": source_id,
            "source_name": source_name,
            "record_status": signal.get("record_status") or run.get("status"),
            "confidence": signal.get("confidence"),
            "evidence_title": signal.get("evidence_title") or item.get("title"),
            "evidence_url": signal.get("evidence_url") or item.get("url"),
            "notes": signal.get("notes"),
            "executed_at": executed_at,
            "_has_candidate": bool(candidate_name),
        }
        for key, value in (item.get("extra") or {}).items():
            if key == "talent_signal":
                continue
            row[key] = value
        rows.append(row)

    if rows:
        return rows

    resolved_status = _resolve_source_status(config=config, run=run, rows=rows)
    placeholder_row = {
        "candidate_name": None,
        "university": None,
        "department": None,
        "email": None,
        "track": _first_track(config),
        "source_id": source_id,
        "source_name": source_name,
        "record_status": "blocked" if resolved_status == "blocked" else "needs_review",
        "confidence": None,
        "evidence_title": _resolve_block_reason(run=run, rows=rows) or source_name,
        "evidence_url": _first_seed_url(config),
        "executed_at": executed_at,
        "_placeholder": True,
    }
    return [placeholder_row]


def _resolve_source_status(
    *,
    config: dict[str, Any],
    run: dict[str, Any],
    rows: list[dict[str, Any]],
) -> str:
    signal_rows = _real_signal_rows(rows)
    if any(row.get("record_status") == "blocked" for row in signal_rows):
        return "blocked"
    if any(row.get("record_status") == "partial" for row in signal_rows):
        return "partial"
    if any(row.get("record_status") == "needs_review" for row in signal_rows):
        return "needs_review"

    status = str(run.get("status") or "")
    if (
        not signal_rows
        and (
            bool(config.get("requires_auth", False))
            or str(config.get("capture_mode") or "") == "evidence_only"
        )
    ):
        return "blocked"
    return status or "pending"


def _resolve_block_reason(*, run: dict[str, Any], rows: list[dict[str, Any]]) -> Any:
    if run.get("error_message"):
        return run.get("error_message")
    signal_rows = _real_signal_rows(rows)
    if any(row.get("record_status") == "blocked" for row in signal_rows):
        return next(
            (
                row.get("notes") or row.get("evidence_title")
                for row in signal_rows
                if row.get("record_status") == "blocked"
                and (row.get("notes") or row.get("evidence_title"))
            ),
            None,
        )
    return None


def _count_exported_rows(rows: list[dict[str, Any]]) -> int:
    return len([row for row in _real_signal_rows(rows) if row.get("_has_candidate", True)])


def _real_signal_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if not row.get("_placeholder")]


def _first_seed_url(config: dict[str, Any]) -> Any:
    seed_urls = config.get("seed_urls") or []
    if isinstance(seed_urls, list):
        for value in seed_urls:
            if isinstance(value, str) and value:
                return value
            if isinstance(value, dict):
                url = value.get("url")
                if url:
                    return url
    return None


def _first_track(config: dict[str, Any]) -> Any:
    tracks = config.get("tracks") or []
    if isinstance(tracks, list) and tracks:
        return tracks[0]
    track = config.get("track")
    return track if track else None


def _format_datetime(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return _normalize_datetime(value).isoformat()
    return str(value)


def _normalize_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)
