"""Scholar creation service — manual scholar creation and Excel bulk import."""
from __future__ import annotations

import csv
import json
import logging
from datetime import UTC, datetime
from io import StringIO
from pathlib import Path
from typing import Any

from app.crawlers.utils.dedup import compute_url_hash
from app.schemas.scholar import ScholarRecord, compute_scholar_completeness
from app.services.scholar._data import _load_all_with_annotations

logger = logging.getLogger(__name__)

# Manual scholars storage path
MANUAL_SCHOLARS_FILE = Path("data/scholars/manual/manual_scholars/latest.json")
MANUAL_SOURCE_ID = "manual_scholars"
MANUAL_SOURCE_NAME = "手动维护学者库"
MANUAL_GROUP = "manual"


def _generate_url_for_scholar(name: str, university: str, department: str, profile_url: str) -> str:
    """Generate a unique URL for a manually created scholar.

    Priority:
    1. If profile_url is provided, use it directly
    2. Otherwise, construct a synthetic manual:// URL from name + university + department
    """
    if profile_url:
        return profile_url

    # Construct synthetic URL: manual://name@university/department
    dept_part = f"/{department}" if department else ""
    return f"manual://{name}@{university}{dept_part}"


def _check_duplicate(name: str, university: str, email: str, phone: str) -> tuple[bool, str]:
    """Check if a scholar with the same name, university, and contact info already exists.

    Duplicate criteria (user requirement):
    - Same name + same university + same contact info (email or phone)
    - If neither has contact info, name + university alone is sufficient

    Returns:
        (is_duplicate, url_hash_of_existing_record)
    """
    all_scholars = _load_all_with_annotations()

    for scholar in all_scholars:
        existing_name = scholar.get("name", "").strip()
        existing_uni = scholar.get("university", "").strip()
        existing_email = scholar.get("email", "").strip()
        existing_phone = scholar.get("phone", "").strip()

        # Name and university must match
        if existing_name.lower() != name.lower():
            continue
        if existing_uni.lower() != university.lower():
            continue

        # Contact info check
        has_contact = bool(email or phone)
        existing_has_contact = bool(existing_email or existing_phone)

        if not has_contact and not existing_has_contact:
            # Neither has contact info — name + university match is enough
            return True, scholar.get("url_hash", "")

        if has_contact and existing_has_contact:
            # Both have contact info — check if any matches
            if email and email.lower() == existing_email.lower():
                return True, scholar.get("url_hash", "")
            if phone and phone == existing_phone:
                return True, scholar.get("url_hash", "")

    return False, ""


def _save_scholar_to_manual_file(item: dict[str, Any]) -> None:
    """Append a new scholar item to the manual scholars JSON file.

    Thread-safe: uses atomic write (write to temp, then replace).
    """
    MANUAL_SCHOLARS_FILE.parent.mkdir(parents=True, exist_ok=True)

    # Load existing data
    if MANUAL_SCHOLARS_FILE.exists():
        with open(MANUAL_SCHOLARS_FILE, encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {
            "source_id": MANUAL_SOURCE_ID,
            "source_name": MANUAL_SOURCE_NAME,
            "group": MANUAL_GROUP,
            "crawled_at": datetime.now(UTC).isoformat(),
            "items": [],
        }

    # Append new item
    data["items"].append(item)
    data["item_count"] = len(data["items"])
    data["crawled_at"] = datetime.now(UTC).isoformat()

    # Atomic write
    tmp_path = MANUAL_SCHOLARS_FILE.with_suffix(".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp_path.replace(MANUAL_SCHOLARS_FILE)


def create_scholar(data: dict[str, Any]) -> tuple[dict[str, Any] | None, str]:
    """Create a new scholar record manually.

    Args:
        data: ScholarCreateRequest fields as dict

    Returns:
        (scholar_detail_dict, error_message)
        - On success: (detail, "")
        - On duplicate: (None, "duplicate:{url_hash}")
        - On error: (None, error_message)
    """
    name = data.get("name", "").strip()
    university = data.get("university", "").strip()
    email = data.get("email", "").strip()
    phone = data.get("phone", "").strip()
    profile_url = data.get("profile_url", "").strip()

    if not name:
        return None, "name is required"

    # Check duplicate
    is_dup, existing_hash = _check_duplicate(name, university, email, phone)
    if is_dup:
        return None, f"duplicate:{existing_hash}"

    # Generate URL and url_hash
    url = _generate_url_for_scholar(name, university, data.get("department", ""), profile_url)
    url_hash = compute_url_hash(url)

    # Build ScholarRecord
    now_iso = datetime.now(UTC).isoformat()
    scholar_record = ScholarRecord(
        name=name,
        name_en=data.get("name_en", ""),
        gender=data.get("gender", ""),
        photo_url=data.get("photo_url", ""),
        university=university,
        department=data.get("department", ""),
        secondary_departments=data.get("secondary_departments", []),
        position=data.get("position", ""),
        academic_titles=data.get("academic_titles", []),
        is_academician=data.get("is_academician", False),
        research_areas=data.get("research_areas", []),
        keywords=data.get("keywords", []),
        bio=data.get("bio", ""),
        bio_en=data.get("bio_en", ""),
        email=email,
        phone=phone,
        office=data.get("office", ""),
        profile_url=profile_url,
        lab_url=data.get("lab_url", ""),
        google_scholar_url=data.get("google_scholar_url", ""),
        dblp_url=data.get("dblp_url", ""),
        orcid=data.get("orcid", ""),
        phd_institution=data.get("phd_institution", ""),
        phd_year=data.get("phd_year", ""),
        education=data.get("education", []),
        source_id=MANUAL_SOURCE_ID,
        source_url="",
        crawled_at=now_iso,
        first_seen_at=now_iso,
        last_seen_at=now_iso,
        is_active=True,
    )

    # Compute completeness
    scholar_record.data_completeness = compute_scholar_completeness(scholar_record)

    # Build CrawledItem structure
    item = {
        "title": name,
        "url": url,
        "url_hash": url_hash,
        "published_at": now_iso,
        "author": "",
        "content": "",
        "content_html": "",
        "content_hash": "",
        "source_id": MANUAL_SOURCE_ID,
        "dimension": "scholars",
        "tags": [],
        "extra": scholar_record.model_dump(),
        "is_new": True,
    }

    # Save to file
    try:
        _save_scholar_to_manual_file(item)
    except Exception as exc:
        logger.error("Failed to save scholar %s: %s", name, exc, exc_info=True)
        return None, f"save failed: {exc}"

    # Return full detail (reload from storage to ensure consistency)
    from app.services.scholar import get_scholar_detail
    detail = get_scholar_detail(url_hash)
    return detail, ""


# ---------------------------------------------------------------------------
# Excel import
# ---------------------------------------------------------------------------

# Column name mapping (case-insensitive, supports Chinese and English)
_COLUMN_MAP = {
    "姓名": "name",
    "name": "name",
    "英文名": "name_en",
    "name_en": "name_en",
    "性别": "gender",
    "gender": "gender",
    "照片": "photo_url",
    "photo": "photo_url",
    "photo_url": "photo_url",
    "高校": "university",
    "大学": "university",
    "university": "university",
    "院系": "department",
    "department": "department",
    "职称": "position",
    "position": "position",
    "学术头衔": "academic_titles",
    "academic_titles": "academic_titles",
    "是否院士": "is_academician",
    "is_academician": "is_academician",
    "研究方向": "research_areas",
    "research_areas": "research_areas",
    "关键词": "keywords",
    "keywords": "keywords",
    "简介": "bio",
    "bio": "bio",
    "邮箱": "email",
    "email": "email",
    "电话": "phone",
    "phone": "phone",
    "办公室": "office",
    "office": "office",
    "个人主页": "profile_url",
    "profile_url": "profile_url",
    "实验室主页": "lab_url",
    "lab_url": "lab_url",
    "google scholar": "google_scholar_url",
    "google_scholar_url": "google_scholar_url",
    "dblp": "dblp_url",
    "dblp_url": "dblp_url",
    "orcid": "orcid",
    "博士院校": "phd_institution",
    "phd_institution": "phd_institution",
    "博士年份": "phd_year",
    "phd_year": "phd_year",
}


def _normalize_column_name(col: str) -> str:
    """Normalize column name to standard field name."""
    col_lower = col.strip().lower()
    return _COLUMN_MAP.get(col_lower, "")


def _parse_bool(value: str) -> bool:
    """Parse boolean from string (是/否/true/false/1/0)."""
    v = value.strip().lower()
    return v in ("是", "true", "1", "yes")


def _parse_list(value: str, delimiter: str = ",") -> list[str]:
    """Parse comma/semicolon-separated list."""
    if not value:
        return []
    # Support both comma and semicolon
    if ";" in value:
        delimiter = ";"
    return [item.strip() for item in value.split(delimiter) if item.strip()]


def _parse_excel_row(row: dict[str, str]) -> dict[str, Any]:
    """Parse a single Excel row into ScholarCreateRequest fields."""
    result: dict[str, Any] = {}

    for col, value in row.items():
        field = _normalize_column_name(col)
        if not field or not value:
            continue

        value = value.strip()

        # Type conversions
        if field == "is_academician":
            result[field] = _parse_bool(value)
        elif field in ("academic_titles", "research_areas", "keywords"):
            result[field] = _parse_list(value)
        else:
            result[field] = value

    return result


def import_scholars_excel(
    file_content: bytes,
    filename: str,
    added_by: str = "user",
    skip_duplicates: bool = True,
) -> dict[str, Any]:
    """Import scholars from Excel/CSV file.

    Args:
        file_content: File bytes
        filename: Original filename (used to detect format)
        added_by: Operator name for audit
        skip_duplicates: If True, skip duplicate scholars; if False, fail on duplicate

    Returns:
        ScholarImportResult dict
    """
    result = {
        "total": 0,
        "success": 0,
        "skipped": 0,
        "failed": 0,
        "items": [],
    }

    # Detect format
    is_csv = filename.lower().endswith(".csv")

    try:
        if is_csv:
            # Parse CSV
            text = file_content.decode("utf-8-sig")  # Handle BOM
            reader = csv.DictReader(StringIO(text))
            rows = list(reader)
        else:
            # Parse Excel (XLSX)
            try:
                import openpyxl
            except ImportError:
                result["failed"] = 1
                result["items"].append({
                    "row": 0,
                    "status": "failed",
                    "name": "",
                    "reason": "openpyxl not installed (run: pip install openpyxl)",
                })
                return result

            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active

            # Read header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip() if cell else "" for cell in header_row]

            # Read data rows
            rows = []
            for row_values in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value).strip() if value else ""
                rows.append(row_dict)

    except Exception as exc:
        logger.error("Failed to parse file %s: %s", filename, exc, exc_info=True)
        result["failed"] = 1
        result["items"].append({
            "row": 0,
            "status": "failed",
            "name": "",
            "reason": f"parse error: {exc}",
        })
        return result

    result["total"] = len(rows)

    # Process each row
    for row_idx, row in enumerate(rows, start=1):
        parsed = _parse_excel_row(row)
        name = parsed.get("name", "").strip()

        if not name:
            result["failed"] += 1
            result["items"].append({
                "row": row_idx,
                "status": "failed",
                "name": "",
                "reason": "name is required",
            })
            continue

        # Add audit field
        parsed["added_by"] = added_by

        # Create scholar
        detail, error = create_scholar(parsed)

        if error.startswith("duplicate:"):
            if skip_duplicates:
                existing_hash = error.split(":", 1)[1]
                result["skipped"] += 1
                result["items"].append({
                    "row": row_idx,
                    "status": "skipped",
                    "name": name,
                    "url_hash": existing_hash,
                    "reason": "duplicate",
                })
            else:
                result["failed"] += 1
                result["items"].append({
                    "row": row_idx,
                    "status": "failed",
                    "name": name,
                    "reason": "duplicate",
                })
        elif error:
            result["failed"] += 1
            result["items"].append({
                "row": row_idx,
                "status": "failed",
                "name": name,
                "reason": error,
            })
        else:
            result["success"] += 1
            result["items"].append({
                "row": row_idx,
                "status": "success",
                "name": name,
                "url_hash": detail["url_hash"] if detail else "",
            })

    return result
